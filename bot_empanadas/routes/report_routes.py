import io
import shutil
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

from flask import make_response, request, session

try:
    from services.pdf_service import generar_pdf_factura
except ImportError:
    from bot_empanadas.services.pdf_service import generar_pdf_factura


def register_report_routes(app, deps: dict):
    db = deps["db"]
    ok = deps["ok"]
    error = deps["error"]
    login_required = deps["login_required"]
    serialize = deps["serialize"]
    send_text_whatsapp = deps.get("send_text_whatsapp")
    normalize_ticket_destination = deps.get("normalize_ticket_destination", lambda value: value)

    def _invoice_documents_dir():
        doc_folder = Path(app.root_path).parent / "documents"
        doc_folder.mkdir(exist_ok=True)
        return doc_folder

    def _invoice_document_filename(folio_factura, extension):
        folio_safe = str(folio_factura or "").strip().upper().replace(" ", "_").replace("/", "_")
        ext = str(extension or "").strip().lower().lstrip(".")
        if not folio_safe or not ext:
            return None
        return f"factura_{folio_safe}.{ext}"

    def _stage_invoice_document(source_path, target_filename=None):
        if not source_path:
            return None
        source = Path(str(source_path))
        if not source.exists() or not source.is_file():
            return None

        doc_folder = _invoice_documents_dir()
        destination = doc_folder / (target_filename or source.name)
        try:
            if source.resolve() != destination.resolve():
                shutil.copy2(str(source), str(destination))
        except FileNotFoundError:
            return None
        return str(destination)

    def _document_public_url(document_path):
        if not document_path:
            return None
        filename = Path(str(document_path)).name
        base_url = (app.config.get("PUBLIC_BASE_URL") or "http://localhost:5000").rstrip("/")
        return f"{base_url}/documents/{filename}"

    def _build_invoice_xml_content(pedido_id, folio_factura, pedido_info, cliente_info, datos_fiscal_info, items):
        pedido = pedido_info if isinstance(pedido_info, dict) else {}
        cliente = cliente_info if isinstance(cliente_info, dict) else {}
        fiscal = datos_fiscal_info if isinstance(datos_fiscal_info, dict) else {}
        rows = items if isinstance(items, list) else []

        lineas_items = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            cantidad = int(item.get("cantidad") or 0)
            precio = float(item.get("precio_unitario") or item.get("precio_unit") or 0)
            subtotal = float(item.get("subtotal") or (cantidad * precio))
            nombre = str(item.get("producto_nombre") or item.get("nombre") or item.get("producto") or "Producto")
            lineas_items.append(
                "    <item><nombre>{}</nombre><cantidad>{}</cantidad><precio_unitario>{:.2f}</precio_unitario><subtotal>{:.2f}</subtotal></item>".format(
                    xml_escape(nombre),
                    cantidad,
                    precio,
                    subtotal,
                )
            )

        return "\n".join(
            [
                '<?xml version="1.0" encoding="UTF-8"?>',
                '<factura_operativa version="1.0">',
                "  <folio>{}</folio>".format(xml_escape(str(folio_factura or ""))),
                "  <pedido_id>{}</pedido_id>".format(int(pedido_id or 0)),
                "  <fecha_emision>{}</fecha_emision>".format(xml_escape(datetime.utcnow().isoformat() + "Z")),
                "  <total>{:.2f}</total>".format(float(pedido.get("total") or 0)),
                "  <cliente>",
                "    <cliente_id>{}</cliente_id>".format(int(cliente.get("cliente_id") or pedido.get("cliente_id") or 0)),
                "    <nombre>{}</nombre>".format(xml_escape(str(cliente.get("nombre") or ""))),
                "    <apellidos>{}</apellidos>".format(xml_escape(str(cliente.get("apellidos") or ""))),
                "    <whatsapp_id>{}</whatsapp_id>".format(xml_escape(str(cliente.get("whatsapp_id") or ""))),
                "  </cliente>",
                "  <fiscal>",
                "    <datos_fiscales_id>{}</datos_fiscales_id>".format(int(fiscal.get("datos_fiscales_id") or fiscal.get("fiscal_id") or 0)),
                "    <rfc>{}</rfc>".format(xml_escape(str(fiscal.get("rfc") or ""))),
                "    <razon_social>{}</razon_social>".format(xml_escape(str(fiscal.get("razon_social") or ""))),
                "    <regimen_fiscal>{}</regimen_fiscal>".format(xml_escape(str(fiscal.get("regimen_fiscal") or ""))),
                "    <uso_cfdi>{}</uso_cfdi>".format(xml_escape(str(fiscal.get("uso_cfdi") or "G01"))),
                "    <email>{}</email>".format(xml_escape(str(fiscal.get("email") or ""))),
                "  </fiscal>",
                "  <items>",
                *lineas_items,
                "  </items>",
                "</factura_operativa>",
            ]
        )

    def _decorate_invoice_preview(preview_data):
        if not isinstance(preview_data, dict):
            return preview_data

        documentos = preview_data.get("documentos") or {}
        folio = preview_data.get("folio_factura")
        for ext in ("pdf", "xml"):
            info = documentos.get(ext)
            if not isinstance(info, dict):
                continue
            path = info.get("path")
            if info.get("ready") and path:
                staged = _stage_invoice_document(path, _invoice_document_filename(folio, ext)) or path
                info["path"] = staged
                info["url"] = _document_public_url(staged)
        preview_data["documentos"] = documentos
        return preview_data

    @login_required(roles=["admin"])
    def api_top_clientes():
        top = db.obtener_top_clientes(limit=20)
        if isinstance(top, dict) and top.get("error"):
            return error(top["error"], 500)
        return ok(top)

    app.add_url_rule("/api/clientes/top20", endpoint="api_top_clientes", view_func=api_top_clientes, methods=["GET"])

    @login_required(roles=["admin"])
    def api_ventas_diarias():
        data = db.obtener_ventas_diarias()
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/ventas/diarias", endpoint="api_ventas_diarias", view_func=api_ventas_diarias, methods=["GET"])

    @login_required(roles=["admin"])
    def api_ventas_mensuales():
        data = db.obtener_ventas_mensuales()
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/ventas/mensuales", endpoint="api_ventas_mensuales", view_func=api_ventas_mensuales, methods=["GET"])

    @login_required(roles=["admin"])
    def api_ventas_anuales():
        data = db.obtener_ventas_anuales()
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/ventas/anuales", endpoint="api_ventas_anuales", view_func=api_ventas_anuales, methods=["GET"])

    @login_required(roles=["admin"])
    def api_kpis_ventas_periodo():
        data = db.obtener_kpis_ventas_periodo()
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/ventas/kpis-periodo", endpoint="api_kpis_ventas_periodo", view_func=api_kpis_ventas_periodo, methods=["GET"])

    @login_required(roles=["admin"])
    def api_reporte_ventas_profesional():
        periodo = (request.args.get("periodo") or "dia").strip().lower()
        fecha_base = (request.args.get("fecha") or "").strip() or None
        busqueda = (request.args.get("q") or request.args.get("buscar") or "").strip() or None

        limit_raw = request.args.get("limit", "300")
        try:
            limit_int = max(1, min(1000, int(limit_raw)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_reporte_ventas_profesional(
            periodo=periodo,
            fecha_base=fecha_base,
            busqueda=busqueda,
            limit=limit_int,
        )
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule(
        "/api/admin/reporte-ventas-profesional",
        endpoint="api_reporte_ventas_profesional",
        view_func=api_reporte_ventas_profesional,
        methods=["GET"],
    )

    @login_required(roles=["admin"])
    def api_reporte_ventas_profesional_xlsx():
        periodo = (request.args.get("periodo") or "dia").strip().lower()
        fecha_base = (request.args.get("fecha") or "").strip() or None
        busqueda = (request.args.get("q") or request.args.get("buscar") or "").strip() or None

        limit_raw = request.args.get("limit", "1000")
        try:
            limit_int = max(1, min(5000, int(limit_raw)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_reporte_ventas_profesional(
            periodo=periodo,
            fecha_base=fecha_base,
            busqueda=busqueda,
            limit=limit_int,
        )
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)

        try:
            from openpyxl import Workbook
        except Exception:
            return error("Falta dependencia openpyxl. Ejecuta: pip install openpyxl", 500)

        payload = serialize(data if isinstance(data, dict) else {})
        resumen = payload.get("resumen") or {}
        rows = payload.get("rows") or []

        wb = Workbook()
        ws_res = wb.active
        if ws_res is None:
            ws_res = wb.create_sheet(title="Resumen")
        else:
            ws_res.title = "Resumen"
        ws_res.append(["Metrica", "Valor"])
        periodo_label = {"dia": "Dia", "semana": "Semana", "mes": "Mes", "ano": "Ano"}.get(
            str(payload.get("periodo") or "dia").strip().lower(),
            str(payload.get("periodo") or "dia"),
        )
        ws_res.append(["Periodo", periodo_label])
        ws_res.append(["Fecha base", payload.get("fecha_base") or ""])
        ws_res.append(["Ventas", resumen.get("ventas") or 0])
        ws_res.append(["Pedidos", resumen.get("pedidos") or 0])
        ws_res.append(["Ticket promedio", resumen.get("ticket_promedio") or 0])
        ws_res.append(["Clientes unicos", resumen.get("clientes_unicos") or 0])
        ws_res.append(["Costo estimado total", resumen.get("costo_estimado_total") or 0])
        ws_res.append(["Utilidad estimada total", resumen.get("utilidad_estimada_total") or 0])
        ws_res.append(["Reserva impuestos pct", resumen.get("reserva_impuestos_pct") or 0])
        ws_res.append(["Reserva impuestos monto", resumen.get("reserva_impuestos_monto") or 0])
        ws_res.append(["Utilidad neta estimada", resumen.get("utilidad_neta_estimada") or 0])
        ws_res.append(["Margen estimado pct", resumen.get("margen_estimado_pct") or 0])
        ws_res.append(["Rapidez preparacion promedio min", resumen.get("rapidez_preparacion_promedio_min") or 0])
        ws_res.append(["Rapidez entrega promedio min", resumen.get("rapidez_entrega_promedio_min") or 0])

        ws_det = wb.create_sheet(title="Detalle")
        ws_det.append([
            "pedido_id",
            "creado_en",
            "cliente",
            "whatsapp_id",
            "metodo_pago",
            "metodo_entrega",
            "estado",
            "productos",
            "piezas",
            "total",
            "costo_estimado",
            "utilidad_estimada",
            "margen_estimado_pct",
            "rapidez_preparacion_min",
            "rapidez_entrega_min",
        ])

        for row in rows:
            if not isinstance(row, dict):
                continue
            ws_det.append([
                row.get("pedido_id"),
                row.get("creado_en"),
                row.get("cliente"),
                row.get("whatsapp_id"),
                row.get("metodo_pago"),
                row.get("metodo_entrega"),
                row.get("estado"),
                row.get("productos"),
                row.get("piezas"),
                row.get("total"),
                row.get("costo_estimado"),
                row.get("utilidad_estimada"),
                row.get("margen_estimado_pct"),
                row.get("rapidez_preparacion_min"),
                row.get("rapidez_entrega_min"),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"reporte_ventas_profesional_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    app.add_url_rule(
        "/api/admin/reporte-ventas-profesional.xlsx",
        endpoint="api_reporte_ventas_profesional_xlsx",
        view_func=api_reporte_ventas_profesional_xlsx,
        methods=["GET"],
    )

    @login_required(roles=["admin"])
    def api_alertas_inventario():
        data = db.obtener_alertas_inventario()
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/inventario/alertas", endpoint="api_alertas_inventario", view_func=api_alertas_inventario, methods=["GET"])

    @login_required(roles=["admin"])
    def api_inventario():
        texto = (request.args.get("q") or request.args.get("texto") or "").strip() or None
        estado_stock = (request.args.get("estado_stock") or "").strip().lower() or None
        proveedor = (request.args.get("proveedor") or "").strip() or None

        limit_raw = request.args.get("limit")
        offset_raw = request.args.get("offset", "0")
        limit_int = None
        offset_int = 0
        if limit_raw not in (None, ""):
            try:
                limit_int = max(1, min(500, int(limit_raw)))
            except ValueError:
                return error("Parametro limit invalido", 400)
        try:
            offset_int = max(0, int(offset_raw))
        except ValueError:
            return error("Parametro offset invalido", 400)

        data = db.obtener_inventario(
            texto=texto,
            estado_stock=estado_stock,
            proveedor=proveedor,
            limit=limit_int,
            offset=offset_int,
        )
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/inventario", endpoint="api_inventario", view_func=api_inventario, methods=["GET"])

    @login_required(roles=["admin"])
    def api_inventario_compras():
        payload = request.get_json(silent=True) or {}

        insumo = payload.get("insumo")
        cantidad = payload.get("cantidad")
        proveedor = payload.get("proveedor")
        costo_total = payload.get("costo_total")
        confirmar_unidad_base = bool(payload.get("confirmar_unidad_base"))

        created = db.registrar_compra_insumo(
            insumo=insumo,
            cantidad=cantidad,
            proveedor=proveedor,
            costo_total=costo_total,
            creado_por=session.get("user", {}).get("username"),
            actor_rol=session.get("user", {}).get("rol", "admin"),
            confirmar_unidad_base=confirmar_unidad_base,
        )
        if isinstance(created, dict) and created.get("error"):
            return error(created["error"], 400)
        return ok(created, 201)

    app.add_url_rule("/api/inventario/compras", endpoint="api_inventario_compras", view_func=api_inventario_compras, methods=["POST"])

    @login_required(roles=["admin"])
    def api_historial_compras_inventario():
        limit = request.args.get("limit", "30")
        try:
            limit_int = max(1, min(200, int(limit)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_compras_insumos(limit=limit_int)
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule(
        "/api/inventario/compras",
        endpoint="api_historial_compras_inventario",
        view_func=api_historial_compras_inventario,
        methods=["GET"],
    )

    @login_required(roles=["admin"])
    def api_admin_resumen_db():
        data = db.obtener_resumen_db()
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule("/api/admin/resumen-db", endpoint="api_admin_resumen_db", view_func=api_admin_resumen_db, methods=["GET"])

    @login_required(roles=["admin"])
    def api_admin_rentabilidad_productos():
        limit = request.args.get("limit", "20")
        try:
            limit_int = max(1, min(200, int(limit)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_rentabilidad_productos(limit=limit_int)
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule(
        "/api/admin/rentabilidad-productos",
        endpoint="api_admin_rentabilidad_productos",
        view_func=api_admin_rentabilidad_productos,
        methods=["GET"],
    )

    @login_required(roles=["admin"])
    def api_admin_rentabilidad_diagnostico():
        limit = request.args.get("limit", "200")
        try:
            limit_int = max(1, min(500, int(limit)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_diagnostico_costos_receta(limit=limit_int)
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)

        if isinstance(data, list):
            normalized = []
            for row in data:
                if not isinstance(row, dict):
                    normalized.append(row)
                    continue

                detalle = row.get("insumos_sin_costo_detalle")
                if isinstance(detalle, list):
                    normalized.append(row)
                    continue

                fallback_names = row.get("insumos_sin_costo")
                if isinstance(fallback_names, list):
                    row["insumos_sin_costo_detalle"] = [
                        {"insumo_id": None, "nombre": str(name)} for name in fallback_names
                    ]
                else:
                    row["insumos_sin_costo_detalle"] = []
                normalized.append(row)
            data = normalized

        return ok(data)

    app.add_url_rule(
        "/api/admin/rentabilidad-diagnostico",
        endpoint="api_admin_rentabilidad_diagnostico",
        view_func=api_admin_rentabilidad_diagnostico,
        methods=["GET"],
    )

    @login_required(roles=["admin"])
    def api_admin_financial_audit():
        fecha_base = (request.args.get("fecha") or "").strip() or None
        limit_raw = request.args.get("limit", "80")
        try:
            limit_int = max(1, min(300, int(limit_raw)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_auditoria_financiera(fecha_base=fecha_base, limit=limit_int)
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    app.add_url_rule(
        "/api/admin/finanzas/auditoria",
        endpoint="api_admin_financial_audit",
        view_func=api_admin_financial_audit,
        methods=["GET"],
    )


    @login_required(roles=["admin"])
    def api_admin_invoice_delivery():
        """POST /api/admin/finanzas/factura
        
        Registra emisión o entrega de factura, con auditoría y notificación al cliente.
        TAMBIÉN genera PDF y lo envía por WhatsApp.
        
        Payload:
        {
            "pedido_id": int,
            "folio_factura": string (required),
            "status": "emitida" | "entregada",
            "notas": string (optional)
        }
        """
        payload = request.get_json(silent=True) or {}
        pedido_raw = payload.get("pedido_id")
        try:
            pedido_id = int(pedido_raw or 0)
        except (TypeError, ValueError):
            return error("Parametro pedido_id invalido", 400)
        
        if pedido_id < 1:
            return error("Pedido debe ser > 0", 400)

        folio_factura = (payload.get("folio_factura") or "").strip().upper()
        if not folio_factura:
            return error("Folio de factura es requerido", 400)
        
        # Validar que folio tiene formato razonable (no muy corto ni muy largo)
        if len(folio_factura) < 3 or len(folio_factura) > 80:
            return error("Folio debe tener 3-80 caracteres", 400)
        
        status = (payload.get("status") or "emitida").strip().lower() or "emitida"
        if status not in {"emitida", "entregada"}:
            return error("Status debe ser 'emitida' o 'entregada'", 400)
        
        notas = (payload.get("notas") or "").strip() or None
        actor_username = session.get("user", {}).get("username", "desconocido")

        # Registrar que se intentó emitir factura
        try:
            db.registrar_auditoria_factura(
                pedido_id=pedido_id,
                evento_tipo="factura_emitida" if status == "emitida" else "factura_entregada",
                detalles={"folio": folio_factura[:4] + "***", "notas_presentes": bool(notas)},
                actor_username=actor_username,
                actor_rol="admin"
            )
        except Exception as e:
            print(f"Advertencia: Error registrando auditoría de emisión: {e}")

        # Registrar factura en BD
        data = db.registrar_factura_operativa(
            pedido_id=pedido_id,
            folio_factura=folio_factura,
            status=status,
            notas=notas,
            actor_usuario=actor_username,
        )
        if isinstance(data, dict) and data.get("error"):
            # Log de error
            try:
                db.registrar_auditoria_factura(
                    pedido_id=pedido_id,
                    evento_tipo="factura_fallida" if status == "emitida" else "factura_entregada_fallida",
                    detalles={"error": data.get("error")},
                    actor_username=actor_username,
                    actor_rol="admin"
                )
            except:
                pass
            return error(data["error"], 400)

        response_data = dict(data or {})
        response_data["notificacion_cliente"] = {
            "enviado": False,
            "motivo": "No se intento enviar notificacion.",
            "pdf_generado": False,
            "xml_generado": False,
        }
        response_data["pdf"] = None
        response_data["xml"] = None

        # GENERAR PDF y XML cuando se marca como emitida/entregada
        pdf_ruta = None
        xml_ruta = None
        if status in {"emitida", "entregada"}:
            try:
                # Obtener datos del pedido para el PDF
                pedido_info = db.obtener_pedido_por_id(pedido_id)
                if not isinstance(pedido_info, dict) or pedido_info.get("error"):
                    print(f"Advertencia: No se pudo obtener datos del pedido {pedido_id}")
                    pedido_info = {}
                
                cliente_info = db.obtener_cliente_por_id(pedido_info.get("cliente_id")) if pedido_info.get("cliente_id") else {}
                if isinstance(cliente_info, dict) and cliente_info.get("error"):
                    cliente_info = {}
                
                datos_fiscales_id = pedido_info.get("datos_fiscales_id") or data.get("datos_fiscales_id")
                datos_fiscal_info = db.obtener_datos_fiscales_por_id(datos_fiscales_id) if datos_fiscales_id else {}
                if isinstance(datos_fiscal_info, dict) and datos_fiscal_info.get("error"):
                    datos_fiscal_info = {}
                
                # Obtener items del pedido
                items = db.obtener_items_pedido(pedido_id)
                if isinstance(items, dict) and items.get("error"):
                    items = []

                # Generar XML operativo en carpeta documents
                try:
                    xml_filename = _invoice_document_filename(folio_factura, "xml")
                    xml_output_path = _invoice_documents_dir() / xml_filename if xml_filename else None
                    if xml_output_path:
                        xml_content = _build_invoice_xml_content(
                            pedido_id=pedido_id,
                            folio_factura=folio_factura,
                            pedido_info=pedido_info,
                            cliente_info=cliente_info,
                            datos_fiscal_info=datos_fiscal_info,
                            items=items,
                        )
                        xml_output_path.write_text(xml_content, encoding="utf-8")
                        if xml_output_path.exists():
                            xml_ruta = str(xml_output_path)
                            response_data["xml"] = {
                                "ruta": xml_ruta,
                                "folio": folio_factura,
                                "generado_en": datetime.utcnow().isoformat() + "Z",
                                "url": _document_public_url(xml_ruta),
                            }
                            response_data["notificacion_cliente"]["xml_generado"] = True
                            try:
                                db.registrar_auditoria_factura(
                                    pedido_id=pedido_id,
                                    evento_tipo="xml_generado",
                                    detalles={"tipo": "xml", "ruta": xml_ruta, "tamano": Path(xml_ruta).stat().st_size},
                                    actor_username=actor_username,
                                    actor_rol="admin",
                                )
                            except Exception:
                                pass
                except Exception as e:
                    response_data["xml_error"] = str(e)
                
                # Generar PDF
                pdf_resultado = generar_pdf_factura(
                    pedido_id=pedido_id,
                    folio_factura=folio_factura,
                    datos_cliente={
                        "nombre": cliente_info.get("nombre", "Cliente"),
                        "apellidos": cliente_info.get("apellidos", ""),
                        "whatsapp_id": cliente_info.get("whatsapp_id", "")
                    },
                    datos_fiscales={
                        "rfc": datos_fiscal_info.get("rfc", "N/A"),
                        "razon_social": datos_fiscal_info.get("razon_social", "N/A"),
                        "regimen_fiscal": datos_fiscal_info.get("regimen_fiscal", "N/A"),
                        "uso_cfdi": datos_fiscal_info.get("uso_cfdi", "G01"),
                        "email": datos_fiscal_info.get("email", "")
                    },
                    items_pedido=items or [],
                    total=float(pedido_info.get("total", 0)),
                    fecha_emision=datetime.fromisoformat(pedido_info.get("creado_en")) if isinstance(pedido_info.get("creado_en"), str) else None,
                    empresa_nombre="QUE CHIMBA",
                    empresa_rfc="QUI123456ABC"  # RFC de ejemplo, cambiar al real
                )
                
                if "error" in pdf_resultado:
                    print(f"⚠️  Error generando PDF: {pdf_resultado['error']}")
                    response_data["pdf_error"] = pdf_resultado["error"]
                else:
                    pdf_ruta = _stage_invoice_document(
                        pdf_resultado.get("ruta"),
                        _invoice_document_filename(folio_factura, "pdf"),
                    ) or pdf_resultado.get("ruta")
                    db.actualizar_documentos_factura(
                        pedido_id=pedido_id,
                        pdf_ruta=pdf_ruta,
                        xml_ruta=xml_ruta,
                    )
                    response_data["pdf"] = {
                        "ruta": pdf_ruta,
                        "folio": pdf_resultado.get("folio"),
                        "generado_en": pdf_resultado.get("fecha_generacion"),
                        "url": _document_public_url(pdf_ruta),
                    }
                    response_data["notificacion_cliente"]["pdf_generado"] = True
                    
                    # Registrar auditoría
                    try:
                        db.registrar_auditoria_factura(
                            pedido_id=pedido_id,
                            evento_tipo="pdf_generado",
                            detalles={"ruta": pdf_ruta, "tamaño": Path(pdf_ruta).stat().st_size if Path(pdf_ruta).exists() else 0},
                            actor_username=actor_username,
                            actor_rol="admin"
                        )
                    except Exception as e:
                        print(f"⚠️  Error registrando PDF en auditoría: {e}")
                    
            except Exception as e:
                print(f"⚠️  Error en proceso de PDF: {e}")
                response_data["pdf_error"] = str(e)

        if not pdf_ruta and xml_ruta:
            db.actualizar_documentos_factura(
                pedido_id=pedido_id,
                pdf_ruta=None,
                xml_ruta=xml_ruta,
            )

        # Enviar notificación + PDF/XML al cliente si factura fue entregada
        if status == "entregada":
            destino = normalize_ticket_destination(response_data.get("whatsapp_id"))
            if destino:
                email_str = (response_data.get("email_destino") or "").strip()

                try:
                    from services.whatsapp_service import send_document_whatsapp
                except ImportError:
                    from bot_empanadas.services.whatsapp_service import send_document_whatsapp
                
                # 1. Enviar PDF si está disponible
                pdf_enviado = False
                if pdf_ruta and Path(pdf_ruta).exists():
                    caption = f"📄 Tu factura #{folio_factura}\nPedido: {pedido_id}"
                    pdf_send_result = send_document_whatsapp(app, destino, pdf_ruta, caption=caption)
                    
                    if isinstance(pdf_send_result, dict) and pdf_send_result.get("ok"):
                        pdf_enviado = True
                        try:
                            db.registrar_auditoria_factura(
                                pedido_id=pedido_id,
                                evento_tipo="pdf_enviado_whatsapp",
                                detalles={"destino": destino, "folio": folio_factura},
                                actor_username=actor_username,
                                actor_rol="admin"
                            )
                        except:
                            pass
                    else:
                        print(f"⚠️  Error enviando PDF por WhatsApp: {pdf_send_result.get('error')}")
                        try:
                            db.registrar_auditoria_factura(
                                pedido_id=pedido_id,
                                evento_tipo="pdf_fallo_whatsapp",
                                detalles={"error": pdf_send_result.get("error"), "destino": destino},
                                actor_username=actor_username,
                                actor_rol="admin"
                            )
                        except:
                            pass

                # 2. Enviar XML si está disponible
                xml_enviado = False
                if xml_ruta and Path(xml_ruta).exists():
                    caption_xml = f"🧾 XML de factura #{folio_factura}\nPedido: {pedido_id}"
                    xml_send_result = send_document_whatsapp(app, destino, xml_ruta, caption=caption_xml)

                    if isinstance(xml_send_result, dict) and xml_send_result.get("ok"):
                        xml_enviado = True
                        try:
                            db.registrar_auditoria_factura(
                                pedido_id=pedido_id,
                                evento_tipo="xml_enviado_whatsapp",
                                detalles={"destino": destino, "folio": folio_factura},
                                actor_username=actor_username,
                                actor_rol="admin"
                            )
                        except:
                            pass
                    else:
                        print(f"⚠️  Error enviando XML por WhatsApp: {xml_send_result.get('error')}")
                        try:
                            db.registrar_auditoria_factura(
                                pedido_id=pedido_id,
                                evento_tipo="xml_fallo_whatsapp",
                                detalles={"error": xml_send_result.get("error"), "destino": destino},
                                actor_username=actor_username,
                                actor_rol="admin"
                            )
                        except:
                            pass
                
                # 3. Enviar mensaje de confirmación por texto
                texto = (
                    f"✅ Tu factura del pedido #{pedido_id} ya fue procesada.\n"
                    f"📋 Folio: {response_data.get('folio_factura')}"
                )
                if pdf_enviado:
                    texto += f"\n📎 PDF adjunto ↑"
                if xml_enviado:
                    texto += f"\n📎 XML adjunto ↑"
                if email_str:
                    texto += f"\n📧 También al correo: {email_str[:30]}..."
                
                enviado = send_text_whatsapp(destino=destino, texto=texto)
                if isinstance(enviado, dict) and enviado.get("error"):
                    response_data["notificacion_cliente"] = {
                        "enviado": False,
                        "motivo": enviado["error"],
                        "destino": destino,
                        "pdf_enviado": pdf_enviado,
                        "xml_enviado": xml_enviado,
                    }
                    # Log de fallo de notificación
                    try:
                        db.registrar_auditoria_factura(
                            pedido_id=pedido_id,
                            evento_tipo="notificacion_whatsapp_fallida",
                            detalles={
                                "error": enviado.get("error"),
                                "destino": destino,
                                "pdf_enviado": pdf_enviado,
                                "xml_enviado": xml_enviado,
                            },
                            actor_username=actor_username,
                            actor_rol="admin"
                        )
                    except:
                        pass
                else:
                    response_data["notificacion_cliente"] = {
                        "enviado": True,
                        "motivo": None,
                        "destino": destino,
                        "pdf_enviado": pdf_enviado,
                        "xml_enviado": xml_enviado,
                    }
                    # Log de éxito de notificación
                    try:
                        db.registrar_auditoria_factura(
                            pedido_id=pedido_id,
                            evento_tipo="notificacion_whatsapp_enviada",
                            detalles={
                                "destino": destino,
                                "pdf_enviado": pdf_enviado,
                                "xml_enviado": xml_enviado,
                            },
                            actor_username=actor_username,
                            actor_rol="admin"
                        )
                    except:
                        pass
                    
                    if hasattr(db, "crear_log_notificacion"):
                        db.crear_log_notificacion({
                            "pedido_id": pedido_id,
                            "canal": "whatsapp",
                            "destino": destino,
                            "tipo": "factura_entregada",
                            "mensaje": texto,
                            "total": response_data.get("total"),
                        })

        return ok(response_data)

    @login_required(roles=["admin"])
    def api_admin_invoice_audit_history():
        """GET /api/admin/finanzas/factura/historial?pedido_id=123
        
        Obtiene historial completo de auditoría de una factura.
        """
        pedido_id_raw = request.args.get("pedido_id", "")
        try:
            pedido_id = int(pedido_id_raw or 0)
        except (TypeError, ValueError):
            return error("Parametro pedido_id invalido", 400)
        
        if pedido_id < 1:
            return error("Pedido debe ser > 0", 400)
        
        historial = db.obtener_historial_factura(pedido_id)
        if isinstance(historial, dict) and historial.get("error"):
            return error(historial.get("error"), 500)
        
        return ok({
            "pedido_id": pedido_id,
            "eventos": list(historial),
            "total": len(historial)
        })

    @login_required(roles=["admin"])
    def api_admin_invoice_panel():
        search = (request.args.get("q") or request.args.get("buscar") or "").strip() or None
        estado = (request.args.get("estado") or "").strip() or None
        envio = (request.args.get("envio") or "").strip() or None
        limit_raw = request.args.get("limit", "200")
        try:
            limit_int = max(1, min(500, int(limit_raw)))
        except ValueError:
            return error("Parametro limit invalido", 400)

        data = db.obtener_panel_facturas(busqueda=search, estado=estado, envio=envio, limit=limit_int)
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 500)
        return ok(data)

    @login_required(roles=["admin"])
    def api_admin_invoice_preview(pedido_id):
        data = db.obtener_preview_factura(pedido_id)
        if isinstance(data, dict) and data.get("error") and hasattr(db, "reparar_factura_pedido"):
            actor_username = session.get("user", {}).get("username", "admin")
            repaired = db.reparar_factura_pedido(pedido_id=pedido_id, actor_usuario=actor_username)
            if not (isinstance(repaired, dict) and repaired.get("error")):
                data = db.obtener_preview_factura(pedido_id)
        if isinstance(data, dict) and data.get("error"):
            return error(data["error"], 404)
        return ok(_decorate_invoice_preview(dict(data)))

    @login_required(roles=["admin"])
    def api_admin_invoice_documents_update(pedido_id):
        payload = request.get_json(silent=True) or {}
        preview = db.obtener_preview_factura(pedido_id)
        if isinstance(preview, dict) and preview.get("error"):
            return error(preview["error"], 404)

        current_docs = (preview.get("documentos") or {}) if isinstance(preview, dict) else {}
        current_pdf = ((current_docs.get("pdf") or {}).get("path") if isinstance(current_docs, dict) else None)
        current_xml = ((current_docs.get("xml") or {}).get("path") if isinstance(current_docs, dict) else None)

        pdf_input = payload.get("pdf_ruta") if "pdf_ruta" in payload else current_pdf
        xml_input = payload.get("xml_ruta") if "xml_ruta" in payload else current_xml

        updated = db.actualizar_documentos_factura(
            pedido_id=pedido_id,
            pdf_ruta=(str(pdf_input).strip() if pdf_input else None),
            xml_ruta=(str(xml_input).strip() if xml_input else None),
        )
        if isinstance(updated, dict) and updated.get("error"):
            return error(updated["error"], 400)

        preview = db.obtener_preview_factura(pedido_id)
        if isinstance(preview, dict) and preview.get("error"):
            return error(preview["error"], 404)
        return ok(_decorate_invoice_preview(dict(preview)))

    @login_required(roles=["admin"])
    def api_admin_invoice_send(pedido_id):
        preview = db.obtener_preview_factura(pedido_id)
        if isinstance(preview, dict) and preview.get("error") and hasattr(db, "reparar_factura_pedido"):
            actor_username = session.get("user", {}).get("username", "admin")
            repaired = db.reparar_factura_pedido(pedido_id=pedido_id, actor_usuario=actor_username)
            if not (isinstance(repaired, dict) and repaired.get("error")):
                preview = db.obtener_preview_factura(pedido_id)
        if isinstance(preview, dict) and preview.get("error"):
            return error(preview["error"], 404)

        preview = _decorate_invoice_preview(dict(preview))
        documentos = preview.get("documentos") or {}
        pdf_info = documentos.get("pdf") or {}
        xml_info = documentos.get("xml") or {}
        destino = normalize_ticket_destination(preview.get("whatsapp_id"))
        actor_username = session.get("user", {}).get("username", "desconocido")

        if not destino:
            return error("La factura no tiene un destino de WhatsApp válido.", 400)
        if not pdf_info.get("ready") or not xml_info.get("ready"):
            return error("Para enviar la factura se requieren PDF y XML listos.", 400)

        pdf_path = _stage_invoice_document(pdf_info.get("path"), _invoice_document_filename(preview.get("folio_factura"), "pdf"))
        xml_path = _stage_invoice_document(xml_info.get("path"), _invoice_document_filename(preview.get("folio_factura"), "xml"))
        if not pdf_path or not xml_path:
            return error("No se pudieron preparar los documentos para envío.", 400)

        db.actualizar_documentos_factura(pedido_id=pedido_id, pdf_ruta=pdf_path, xml_ruta=xml_path)

        try:
            from services.whatsapp_service import send_document_whatsapp
        except ImportError:
            from bot_empanadas.services.whatsapp_service import send_document_whatsapp

        caption_pdf = f"Factura {preview.get('folio_factura')} · Pedido #{pedido_id}"
        caption_xml = f"XML CFDI {preview.get('folio_factura')} · Pedido #{pedido_id}"
        pdf_result = send_document_whatsapp(app, destino, pdf_path, caption=caption_pdf)
        if isinstance(pdf_result, dict) and pdf_result.get("error"):
            db.registrar_resultado_envio_factura(pedido_id, "error", destino=destino, error_detalle=pdf_result.get("error"))
            db.registrar_auditoria_factura(
                pedido_id=pedido_id,
                evento_tipo="notificacion_whatsapp_fallida",
                detalles={"fase": "pdf", "error": pdf_result.get("error"), "destino": destino},
                actor_username=actor_username,
                actor_rol="admin",
            )
            return error(pdf_result.get("error") or "No se pudo enviar el PDF.", 502)

        xml_result = send_document_whatsapp(app, destino, xml_path, caption=caption_xml)
        if isinstance(xml_result, dict) and xml_result.get("error"):
            db.registrar_resultado_envio_factura(pedido_id, "error", destino=destino, error_detalle=xml_result.get("error"))
            db.registrar_auditoria_factura(
                pedido_id=pedido_id,
                evento_tipo="notificacion_whatsapp_fallida",
                detalles={"fase": "xml", "error": xml_result.get("error"), "destino": destino},
                actor_username=actor_username,
                actor_rol="admin",
            )
            return error(xml_result.get("error") or "No se pudo enviar el XML.", 502)

        text_warning = None
        if send_text_whatsapp:
            texto = (
                f"✅ Tu factura #{preview.get('folio_factura')} ya fue enviada.\n"
                f"Pedido: #{pedido_id}\n"
                f"Incluye PDF y XML."
            )
            text_result = send_text_whatsapp(destino=destino, texto=texto)
            if isinstance(text_result, dict) and text_result.get("error"):
                text_warning = text_result.get("error")

        db.registrar_resultado_envio_factura(pedido_id, "enviado", destino=destino, error_detalle=text_warning, marcar_entregada=True)
        db.registrar_auditoria_factura(
            pedido_id=pedido_id,
            evento_tipo="notificacion_whatsapp_enviada",
            detalles={"destino": destino, "documentos": [Path(pdf_path).name, Path(xml_path).name], "warning": text_warning},
            actor_username=actor_username,
            actor_rol="admin",
        )

        updated_preview = db.obtener_preview_factura(pedido_id)
        if isinstance(updated_preview, dict) and updated_preview.get("error"):
            return error(updated_preview["error"], 404)
        return ok(_decorate_invoice_preview(dict(updated_preview)))

    @login_required(roles=["admin"])
    def api_admin_invoice_repair(pedido_id):
        actor_username = session.get("user", {}).get("username", "admin")
        repaired = db.reparar_factura_pedido(pedido_id=pedido_id, actor_usuario=actor_username)
        if isinstance(repaired, dict) and repaired.get("error"):
            return error(repaired["error"], 400)

        preview = db.obtener_preview_factura(pedido_id)
        if isinstance(preview, dict) and preview.get("error"):
            return error(preview["error"], 404)
        payload = dict(preview)
        payload["repair_result"] = repaired
        return ok(_decorate_invoice_preview(payload))

    app.add_url_rule(
        "/api/admin/finanzas/factura",
        endpoint="api_admin_invoice_delivery",
        view_func=api_admin_invoice_delivery,
        methods=["POST"],
    )

    app.add_url_rule(
        "/api/admin/finanzas/factura/historial",
        endpoint="api_admin_invoice_audit_history",
        view_func=api_admin_invoice_audit_history,
        methods=["GET"],
    )

    app.add_url_rule(
        "/api/admin/finanzas/facturas",
        endpoint="api_admin_invoice_panel",
        view_func=api_admin_invoice_panel,
        methods=["GET"],
    )

    app.add_url_rule(
        "/api/admin/finanzas/facturas/<int:pedido_id>/preview",
        endpoint="api_admin_invoice_preview",
        view_func=api_admin_invoice_preview,
        methods=["GET"],
    )

    app.add_url_rule(
        "/api/admin/finanzas/facturas/<int:pedido_id>/documentos",
        endpoint="api_admin_invoice_documents_update",
        view_func=api_admin_invoice_documents_update,
        methods=["PATCH"],
    )

    app.add_url_rule(
        "/api/admin/finanzas/facturas/<int:pedido_id>/send",
        endpoint="api_admin_invoice_send",
        view_func=api_admin_invoice_send,
        methods=["POST"],
    )

    app.add_url_rule(
        "/api/admin/finanzas/facturas/<int:pedido_id>/repair",
        endpoint="api_admin_invoice_repair",
        view_func=api_admin_invoice_repair,
        methods=["POST"],
    )

    @login_required(roles=["admin"])
    def serve_document(filename):
        """Sirve un documento (PDF) desde la carpeta de documentos.
        
        Ruta: /documents/<filename>
        """
        from flask import send_from_directory
        import os
        
        # Validar que el nombre del archivo no contiene path traversal
        if ".." in filename or "/" in filename or "\\" in filename:
            return error("Nombre de archivo inválido", 400)
        
        # Ruta base de documentos (relativos a la carpeta de la app)
        doc_folder = Path(app.root_path).parent / "documents"
        doc_folder.mkdir(exist_ok=True)
        
        document_path = doc_folder / filename
        
        # Verificar que el archivo existe y está dentro de doc_folder
        if not document_path.exists() or not str(document_path).startswith(str(doc_folder)):
            return error("Documento no encontrado", 404)
        
        return send_from_directory(str(doc_folder), filename)

    app.add_url_rule(
        "/documents/<filename>",
        endpoint="serve_document",
        view_func=serve_document,
        methods=["GET"],
    )
