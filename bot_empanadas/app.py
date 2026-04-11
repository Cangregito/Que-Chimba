import json
import importlib
import logging
import os
import csv
import io
import secrets
from datetime import date, datetime
from decimal import Decimal
from functools import wraps
from typing import Any, cast
from urllib.parse import quote
from xml.sax.saxutils import escape as xml_escape

import requests
from flask import Flask, abort, jsonify, redirect, render_template, request, send_from_directory, session, url_for, make_response
from jinja2 import BaseLoader, ChoiceLoader, FileSystemLoader

class _FallbackMessage:
    def __init__(self):
        self._body = ""
        self._media = []

    def body(self, text):
        self._body = str(text)

    def media(self, url):
        self._media.append(str(url))

    def _to_xml(self):
        media_xml = "".join(f"<Media>{xml_escape(m)}</Media>" for m in self._media)
        body_xml = f"<Body>{xml_escape(self._body)}</Body>" if self._body else ""
        return f"<Message>{body_xml}{media_xml}</Message>"


class MessagingResponse:
    def __init__(self):
        self._messages = []

    def message(self):
        msg = _FallbackMessage()
        self._messages.append(msg)
        return msg

    def __str__(self):
        return "<Response>" + "".join(m._to_xml() for m in self._messages) + "</Response>"

import db

try:
    _bot_module = importlib.import_module("bot")
except Exception:
    try:
        _bot_module = importlib.import_module("bot_empanadas.bot")
    except Exception:
        _bot_module = None

if _bot_module is not None:
    procesar_mensaje_whatsapp = _bot_module.procesar_mensaje_whatsapp
else:
    def procesar_mensaje_whatsapp(
        whatsapp_id,
        mensaje,
        media_url=None,
        media_type=None,
        latitude=None,
        longitude=None,
    ):
        texto = (
            f"Ay que chimba, {whatsapp_id}. Recibi: '{mensaje or 'audio'}'. "
            "Listo parce, tu pedido va en camino en el flujo demo."
        )
        return {"tipo": "texto", "contenido": texto}


app = Flask(__name__, template_folder=os.path.join(os.path.dirname(__file__), "templates"))
_flask_secret = (os.getenv("FLASK_SECRET", "") or "").strip()
if not _flask_secret:
    _flask_secret = secrets.token_hex(32)
app.config["SECRET_KEY"] = _flask_secret
app.config["JSON_SORT_KEYS"] = False
app.config["AUDIO_DIR"] = os.path.join(os.path.dirname(__file__), "audios_temp")
app.config["IMG_DIR"] = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Img"))
app.config["PUBLIC_BASE_URL"] = os.getenv("PUBLIC_BASE_URL", "").rstrip("/")
app.config["N8N_PEDIDO_WEBHOOK_URL"] = os.getenv("N8N_PEDIDO_WEBHOOK_URL", "").strip()
app.config["BAILEYS_BRIDGE_URL"] = (os.getenv("BAILEYS_BRIDGE_URL", "http://localhost:3001") or "http://localhost:3001").rstrip("/")
app.config["BAILEYS_BRIDGE_API_TOKEN"] = (os.getenv("BAILEYS_BRIDGE_API_TOKEN", "") or "").strip()
app.config["BAILEYS_WEBHOOK_TOKEN"] = (os.getenv("BAILEYS_WEBHOOK_TOKEN", "") or "").strip()
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = (os.getenv("SESSION_COOKIE_SAMESITE", "Lax") or "Lax").strip()
app.config["SESSION_COOKIE_SECURE"] = (os.getenv("SESSION_COOKIE_SECURE", "0") or "0").strip().lower() in {"1", "true", "yes", "on"}

root_templates_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "templates"))
if os.path.isdir(root_templates_dir):
    jinja_loaders: list[BaseLoader] = [FileSystemLoader(root_templates_dir)]
    if app.jinja_loader is not None:
        jinja_loaders.insert(0, app.jinja_loader)
    app.jinja_loader = ChoiceLoader(jinja_loaders)  # type: ignore[assignment]

os.makedirs(app.config["AUDIO_DIR"], exist_ok=True)

_log_level_name = (os.getenv("LOG_LEVEL", "INFO") or "INFO").upper()
_log_level = getattr(logging, _log_level_name, logging.INFO)
if not logging.getLogger().handlers:
    logging.basicConfig(
        level=_log_level,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    )
app.logger.setLevel(_log_level)

if not (os.getenv("FLASK_SECRET", "") or "").strip():
    app.logger.warning("FLASK_SECRET no esta configurado. Se usa una clave aleatoria temporal para esta ejecucion.")


ESTADOS_VALIDOS_PEDIDO = {
    "recibido",
    "en_preparacion",
    "listo",
    "en_camino",
    "entregado",
    "cancelado",
}


def _json_default(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def _serialize(value):
    return json.loads(json.dumps(value, default=_json_default))


def _ok(data, status=200):
    return jsonify({"ok": True, "data": _serialize(data)}), status


def _generar_error_id() -> str:
    return f"ERR-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(3).upper()}"


def _espera_json() -> bool:
    path = request.path or ""
    if path.startswith("/api/"):
        return True
    accept = (request.headers.get("Accept") or "").lower()
    return "application/json" in accept


def _error(message, status=400, code=None, details=None, error_id=None):
    resolved_error_id = error_id or _generar_error_id()
    payload = {
        "ok": False,
        "error": str(message),
        "error_id": resolved_error_id,
    }
    if code:
        payload["code"] = str(code)
    if details is not None:
        payload["details"] = _serialize(details)

    response = jsonify(payload)
    response.status_code = status
    response.headers["X-Error-ID"] = resolved_error_id
    return response


def _parse_json_field(value, fallback):
    if value is None:
        return fallback
    if isinstance(value, (list, dict)):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return fallback
        return json.loads(raw)
    return fallback


def _validar_requeridos(payload, required_fields):
    missing = []
    for field in required_fields:
        value = payload.get(field)
        if value is None:
            missing.append(field)
            continue
        if isinstance(value, str) and not value.strip():
            missing.append(field)
    return missing


def _normalizar_whatsapp_id(raw_value):
    raw = (raw_value or "").strip()
    if raw.startswith("whatsapp:"):
        raw = raw.replace("whatsapp:", "", 1)
    return raw


def _normalizar_destino_ticket_whatsapp(raw_value):
    raw = _normalizar_whatsapp_id(raw_value)
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    if not digits:
        return ""
    if len(digits) == 10:
        return f"52{digits}"
    return digits


def _client_ip():
    forwarded = request.headers.get("X-Forwarded-For", "").strip()
    if forwarded:
        return forwarded.split(",")[0].strip()
    return (request.remote_addr or "").strip() or None


def _es_origen_valido() -> bool:
    expected = request.host_url.rstrip("/").lower()
    origin = (request.headers.get("Origin") or "").strip().lower()
    referer = (request.headers.get("Referer") or "").strip().lower()

    if origin:
        return origin.startswith(expected)
    if referer:
        return referer.startswith(expected)
    return True


@app.before_request
def _hardening_before_request():
    method = request.method.upper()
    if method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None

    path = request.path or ""
    if path.startswith("/webhook") or path in {"/login"}:
        return None

    if path.startswith("/api/") and session.get("user"):
        if not _es_origen_valido():
            return _error("Origen no autorizado", 403)

    return None


def _resumen_items_para_alerta(items):
    if not isinstance(items, list) or not items:
        return "sin productos"

    parts = []
    for item in items:
        if not isinstance(item, dict):
            continue
        producto = item.get("producto") or item.get("nombre") or f"producto_id={item.get('producto_id', '?')}"
        cantidad = item.get("cantidad", 1)
        parts.append(f"{producto} x{cantidad}")

    return ", ".join(parts) if parts else "sin productos"


def _notificar_n8n_nuevo_pedido(created, payload):
    webhook_url = app.config.get("N8N_PEDIDO_WEBHOOK_URL", "")
    if not webhook_url:
        return

    body = {
        "id": created.get("pedido_id"),
        "productos": payload.get("productos") or _resumen_items_para_alerta(payload.get("items", [])),
        "total": float(created.get("total", 0) or 0),
        "direccion": payload.get("direccion") or payload.get("direccion_texto") or f"direccion_id={payload.get('direccion_id', 'N/A')}",
        "tipo": payload.get("tipo", "individual"),
    }

    try:
        requests.post(webhook_url, json=body, timeout=5)
    except Exception as exc:
        app.logger.warning("No se pudo notificar a n8n para pedido %s: %s", body["id"], exc)


def _enviar_texto_whatsapp(destino, texto):
    bridge_url = app.config.get("BAILEYS_BRIDGE_URL", "")
    if not bridge_url:
        return {"error": "BAILEYS_BRIDGE_URL no configurado."}

    bridge_token = app.config.get("BAILEYS_BRIDGE_API_TOKEN", "")
    headers = {"Content-Type": "application/json"}
    if bridge_token:
        headers["x-bridge-token"] = bridge_token

    try:
        resp = requests.post(
            f"{bridge_url}/api/send-text",
            json={"to": destino, "text": texto},
            timeout=10,
            headers=headers,
        )
    except Exception as exc:
        return {"error": f"No se pudo conectar al bridge de WhatsApp: {exc}"}

    try:
        payload = resp.json()
    except Exception:
        payload = {}

    if not resp.ok or payload.get("ok") is not True:
        msg = payload.get("error") or f"Bridge respondio HTTP {resp.status_code}"
        return {"error": str(msg)}

    return {"ok": True}


def _enviar_audio_whatsapp(destino, audio_path, caption=""):
    """Envia una nota de voz al usuario via el bridge de Baileys."""
    bridge_url = app.config.get("BAILEYS_BRIDGE_URL", "")
    if not bridge_url or not audio_path:
        return {"error": "BAILEYS_BRIDGE_URL o audio_path no configurados."}

    bridge_token = app.config.get("BAILEYS_BRIDGE_API_TOKEN", "")
    headers = {"Content-Type": "application/json"}
    if bridge_token:
        headers["x-bridge-token"] = bridge_token

    # Convertir ruta local a URL publica para que el bridge la descargue.
    audio_filename = os.path.basename(str(audio_path))
    base_url = app.config.get("PUBLIC_BASE_URL") or "http://localhost:5000"
    audio_url = f"{base_url}/audio/{audio_filename}"

    try:
        resp = requests.post(
            f"{bridge_url}/api/send-audio",
            json={"to": destino, "audioUrl": audio_url, "caption": caption},
            timeout=10,
            headers=headers,
        )
    except Exception as exc:
        return {"error": f"No se pudo enviar audio al bridge: {exc}"}

    try:
        payload = resp.json()
    except Exception:
        payload = {}

    if not resp.ok or payload.get("ok") is not True:
        msg = payload.get("error") or f"Bridge respondio HTTP {resp.status_code}"
        return {"error": str(msg)}

    return {"ok": True}


def login_required(roles=None):
    roles = roles or []

    def decorator(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            usuario = session.get("user")
            if not usuario:
                if _espera_json():
                    return _error("Sesion requerida", 401, code="auth_required")
                return redirect(url_for("login_page", next=request.path))

            if roles and usuario.get("rol") not in roles:
                if _espera_json():
                    return _error("No autorizado para esta accion", 403, code="forbidden_role")
                abort(403)
            return view_func(*args, **kwargs)

        return wrapper

    return decorator


@app.get("/health")
def health():
    return _ok({"status": "up"})


@app.get("/img/<path:filename>")
def serve_img(filename):
    img_dir = app.config.get("IMG_DIR", "")
    if not img_dir or not os.path.isdir(img_dir):
        abort(404)
    return send_from_directory(img_dir, filename)


@app.get("/Img/<path:filename>")
def serve_img_upper(filename):
    return serve_img(filename)


@app.get("/favicon.ico")
def favicon():
    return serve_img("simbolo-cuadrado-amarillo.png")


@app.get("/")
def landing():
    raw_number = os.getenv("WHATSAPP_PUBLIC_NUMBER", "526567751166")
    message = os.getenv("WHATSAPP_PUBLIC_MESSAGE", "Hola Que Chimba, quiero pedir unas empanadas")

    wa_number = "".join(ch for ch in str(raw_number) if ch.isdigit())
    whatsapp_url = f"https://wa.me/{wa_number}?text={quote(message)}"
    whatsapp_display = f"+{wa_number}" if wa_number else "WhatsApp"

    return render_template(
        "index.html",
        whatsapp_url=whatsapp_url,
        whatsapp_display=whatsapp_display,
    )


@app.get("/login")
def login_page():
    return render_template("login.html")


@app.get("/soporte")
def soporte_page():
    return render_template("soporte.html")


@app.get("/admin/tickets")
@login_required(roles=["admin"])
def admin_tickets_page():
    return render_template("tickets_admin.html", user=session.get("user"))


@app.post("/login")
def login_post():
    payload = request.get_json(silent=True) or request.form
    username = payload.get("username")
    password = payload.get("password")

    if not isinstance(username, str) or not isinstance(password, str):
        return _error("Credenciales invalidas", 401, code="invalid_credentials")

    username = username.strip()
    if not username or not password:
        return _error(
            "Completa usuario y contrasena",
            400,
            code="validation_error",
            details={"fields": ["username", "password"]},
        )

    user = db.autenticar_usuario(username=username, password=password, direccion_ip=_client_ip())
    if isinstance(user, dict) and user.get("error"):
        return _error(user["error"], int(user.get("status", 401)), code="auth_failed")

    user_session = {
        "usuario_id": user.get("usuario_id"),
        "username": user.get("username"),
        "rol": user.get("rol"),
        "nombre_mostrar": user.get("nombre_mostrar"),
        "area_entrega": user.get("area_entrega"),
    }
    cast(Any, session)["user"] = user_session
    return _ok(user_session)


@app.post("/logout")
def logout():
    if session.get("user") and not _es_origen_valido():
        return _error("Origen no autorizado", 403, code="forbidden_origin")

    actor = session.get("user", {})
    if actor:
        db.registrar_evento_seguridad(
            tipo_evento="logout_success",
            severidad="info",
            actor_usuario_id=actor.get("usuario_id"),
            actor_username=actor.get("username"),
            actor_rol=actor.get("rol"),
            direccion_ip=_client_ip(),
        )
    session.clear()
    return _ok({"message": "Sesion cerrada"})


@app.get("/admin")
@login_required(roles=["admin"])
def admin_page():
    return render_template("admin.html", user=session.get("user"))


@app.get("/cocina")
@login_required(roles=["admin", "cocina"])
def cocina_page():
    return render_template("cocina.html", user=session.get("user"))


@app.get("/repartidor")
@login_required(roles=["admin", "repartidor"])
def repartidor_page():
    return render_template("repartidor.html", user=session.get("user"))


@app.get("/audio/<path:filename>")
def servir_audio(filename):
    return send_from_directory(app.config["AUDIO_DIR"], filename, mimetype="audio/ogg")


@app.get("/api/productos")
def api_productos():
    solo_pedibles_raw = (request.args.get("solo_pedibles") or "1").strip().lower()
    solo_pedibles = solo_pedibles_raw not in {"0", "false", "no", "off"}
    data = db.obtener_productos(solo_pedibles=solo_pedibles)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/stats/publicas")
def api_stats_publicas():
    """Endpoint publico compatible con la landing antigua."""
    resumen = db.obtener_resumen_db()
    if isinstance(resumen, dict) and resumen.get("error"):
        # Mantiene la landing funcional aunque no haya DB disponible.
        return jsonify(
            {
                "total_empanadas_vendidas": 0,
                "total_clientes_felices": 0,
                "total_eventos": 0,
                "source": "fallback",
            }
        )

    total_empanadas = int(resumen.get("pedidos", 0) or 0)
    total_clientes = int(resumen.get("clientes", 0) or 0)
    total_eventos = int(resumen.get("campanas", 0) or 0)

    return jsonify(
        {
            "total_empanadas_vendidas": total_empanadas,
            "total_clientes_felices": total_clientes,
            "total_eventos": total_eventos,
        }
    )


@app.get("/api/evaluaciones/publicas")
def api_evaluaciones_publicas():
    """Endpoint publico para testimonios; tolera esquemas sin tabla de evaluaciones."""
    conn = None
    try:
        conn = db.get_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND table_name = 'evaluaciones'
                LIMIT 1
                """
            )
            if not cur.fetchone():
                return jsonify([])

            cur.execute(
                """
                SELECT
                    e.calificacion,
                    COALESCE(NULLIF(TRIM(e.comentario), ''), 'Excelente servicio y sabor.') AS comentario,
                    COALESCE(NULLIF(TRIM(c.nombre), ''), 'Cliente') AS nombre_cliente
                FROM evaluaciones e
                LEFT JOIN pedidos p ON p.pedido_id = e.pedido_id
                LEFT JOIN clientes c ON c.cliente_id = p.cliente_id
                WHERE e.calificacion IS NOT NULL
                ORDER BY COALESCE(e.creado_en, NOW()) DESC
                LIMIT 12
                """
            )
            rows = cur.fetchall() or []

        payload = []
        for row in rows:
            if isinstance(row, dict):
                payload.append(
                    {
                        "calificacion": int(row.get("calificacion") or 0),
                        "comentario": row.get("comentario") or "Excelente servicio y sabor.",
                        "nombre_cliente": row.get("nombre_cliente") or "Cliente",
                    }
                )
                continue

            payload.append(
                {
                    "calificacion": int(row[0] or 0),
                    "comentario": row[1] or "Excelente servicio y sabor.",
                    "nombre_cliente": row[2] or "Cliente",
                }
            )

        return jsonify(payload)
    except Exception:
        return jsonify([])
    finally:
        if conn:
            conn.close()


@app.get("/api/pedidos")
@login_required(roles=["admin", "cocina", "repartidor"])
def api_pedidos():
    estado_raw = request.args.get("estado")
    estado = None
    if estado_raw:
        estados = [part.strip() for part in estado_raw.split(",") if part.strip()]
        estado = estados if len(estados) > 1 else estados[0]
    fecha = request.args.get("fecha")
    fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
    busqueda = (request.args.get("q") or request.args.get("buscar") or "").strip() or None

    limit_raw = request.args.get("limit")
    offset_raw = request.args.get("offset", "0")
    limit_int = None
    offset_int = 0
    if limit_raw not in (None, ""):
        try:
            limit_int = max(1, min(500, int(limit_raw)))
        except ValueError:
            return _error("Parametro limit invalido", 400)
    try:
        offset_int = max(0, int(offset_raw))
    except ValueError:
        return _error("Parametro offset invalido", 400)

    data = db.obtener_pedidos(
        estado=estado,
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        busqueda=busqueda,
        limit=limit_int,
        offset=offset_int,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.post("/api/pedidos")
def api_crear_pedido():
    payload = request.get_json(silent=True) or {}
    actor = session.get("user", {})

    cliente_id = payload.get("cliente_id")
    whatsapp_id = payload.get("whatsapp_id")
    direccion_id = payload.get("direccion_id")
    metodo_pago = payload.get("metodo_pago", "efectivo")
    items = payload.get("items", [])

    if not cliente_id:
        if not whatsapp_id:
            return _error(
                "Debes enviar cliente_id o whatsapp_id",
                400,
                code="validation_error",
                details={"fields": ["cliente_id", "whatsapp_id"]},
            )
        cliente = db.obtener_o_crear_cliente(whatsapp_id)
        if isinstance(cliente, dict) and cliente.get("error"):
            return _error(cliente["error"], 500)
        cliente_id = cliente["cliente_id"]

    if not isinstance(items, list) or not items:
        return _error(
            "Debes enviar al menos un item en el pedido",
            400,
            code="validation_error",
            details={"fields": ["items"]},
        )

    created = db.crear_pedido(
        cliente_id,
        items,
        direccion_id,
        metodo_pago,
        actor_usuario=actor.get("username") or whatsapp_id or "bot_whatsapp",
        actor_rol=actor.get("rol") or "cliente",
    )
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 500)

    _notificar_n8n_nuevo_pedido(created, payload)
    return _ok(created, 201)


@app.post("/api/logs")
def api_crear_log_notificacion():
    payload = request.get_json(silent=True) or {}

    required = ["pedido_id", "canal", "destino", "tipo", "mensaje"]
    faltantes = _validar_requeridos(payload, required)
    if faltantes:
        return _error(
            f"Campos obligatorios faltantes: {', '.join(faltantes)}",
            400,
            code="validation_error",
            details={"fields": faltantes},
        )

    created = db.crear_log_notificacion(payload)
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 500)

    return _ok(created, 201)


@app.patch("/api/pedidos/<int:pedido_id>/estado")
@login_required(roles=["admin", "cocina", "repartidor"])
def api_actualizar_estado_pedido(pedido_id):
    payload = request.get_json(silent=True) or {}
    nuevo_estado = payload.get("estado")
    motivo = payload.get("motivo")

    if nuevo_estado not in ESTADOS_VALIDOS_PEDIDO:
        return _error(
            "Estado no valido",
            400,
            code="validation_error",
            details={"fields": ["estado"], "allowed": sorted(list(ESTADOS_VALIDOS_PEDIDO))},
        )

    actor = session.get("user", {})
    updated = db.actualizar_estado_pedido(
        pedido_id,
        nuevo_estado,
        actor_usuario=actor.get("username", "sistema"),
        rol_actor=actor.get("rol", "sistema"),
        motivo=motivo,
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        if "no encontrado" in msg:
            status = 404
        elif "transicion no permitida" in msg or "estado no valido" in msg:
            status = 400
        else:
            status = 500
        return _error(updated["error"], status)
    return _ok(updated)


@app.get("/api/repartidor/pedidos")
@login_required(roles=["admin", "repartidor"])
def api_repartidor_pedidos():
    user = session.get("user", {})
    repartidor_usuario = None if user.get("rol") == "admin" else user.get("username")
    area_entrega = None if user.get("rol") == "admin" else user.get("area_entrega")

    data = db.obtener_pedidos_repartidor(
        repartidor_usuario=repartidor_usuario,
        area_entrega=area_entrega,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)

    pedidos = []
    rows = data if isinstance(data, list) else []
    for row in rows:
        if not isinstance(row, dict):
            continue
        productos = []
        for item in row.get("items") or []:
            if not isinstance(item, dict):
                continue
            nombre = item.get("producto") or item.get("nombre") or "Producto"
            variante = item.get("variante") or ""
            cantidad = item.get("cantidad") or 1
            label = f"{cantidad} x {nombre} {variante}".strip()
            productos.append(label)

        pedidos.append(
            {
                "pedido_id": row.get("pedido_id"),
                "estado": row.get("estado"),
                "metodo_pago": row.get("metodo_pago") or "efectivo",
                "cliente_nombre": " ".join(
                    [part for part in [row.get("nombre"), row.get("apellidos")] if part]
                ).strip()
                or "Cliente",
                "direccion_entrega": row.get("direccion_entrega") or "Sin direccion",
                "codigo_postal": row.get("codigo_postal") or "00000",
                "area_entrega": row.get("area_entrega") or row.get("codigo_postal") or "N/A",
                "productos": productos,
            }
        )

    return _ok(pedidos)


@app.post("/api/pedidos/<int:pedido_id>/confirmar")
@login_required(roles=["admin", "repartidor"])
def api_confirmar_entrega_pedido(pedido_id):
    payload = request.get_json(silent=True) or {}
    codigo_entrega = payload.get("codigo_entrega")
    numero_confirmacion_pago = payload.get("numero_confirmacion_pago")
    actor = session.get("user", {})

    updated = db.confirmar_entrega_pedido(
        pedido_id=pedido_id,
        codigo_entrega=codigo_entrega,
        numero_confirmacion_pago=numero_confirmacion_pago,
        actor_usuario=actor.get("username"),
        rol_actor=actor.get("rol"),
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(updated["error"], status)

    response_data: dict[str, Any] = dict(updated or {})
    response_data["notificacion_cliente"] = {
        "enviado": False,
        "motivo": "No se intento enviar notificacion.",
    }

    destino_data = db.obtener_destino_whatsapp_por_pedido(pedido_id=pedido_id)
    if isinstance(destino_data, dict) and not destino_data.get("error"):
        destino = _normalizar_whatsapp_id(destino_data.get("whatsapp_id"))
        confirmacion_pago = (response_data.get("confirmacion_pago") or "").strip()
        mensaje = (
            f"Muchisimas gracias por tu compra, parce. Confirmamos que tu pedido #{pedido_id} "
            "ya fue entregado. Que lo disfrutes."
        )
        if confirmacion_pago:
            mensaje = (
                f"{mensaje} Confirmacion de pago: {confirmacion_pago}."
            )
        enviado = _enviar_texto_whatsapp(destino=destino, texto=mensaje)

        if isinstance(enviado, dict) and enviado.get("error"):
            app.logger.warning(
                "Pedido %s liberado, pero no se pudo enviar agradecimiento al cliente (%s): %s",
                pedido_id,
                destino,
                enviado["error"],
            )
            response_data["notificacion_cliente"] = {
                "enviado": False,
                "motivo": enviado["error"],
                "destino": destino,
            }
        else:
            db.crear_log_notificacion(
                {
                    "pedido_id": pedido_id,
                    "canal": "whatsapp",
                    "destino": destino,
                    "tipo": "agradecimiento_entrega",
                    "mensaje": mensaje,
                    "total": None,
                    "direccion": None,
                }
            )
            response_data["notificacion_cliente"] = {
                "enviado": True,
                "destino": destino,
            }
    else:
        motivo = destino_data.get("error") if isinstance(destino_data, dict) else "Sin detalle."
        app.logger.warning(
            "Pedido %s liberado, pero no se encontro WhatsApp destino para agradecimiento: %s",
            pedido_id,
            motivo,
        )
        response_data["notificacion_cliente"] = {
            "enviado": False,
            "motivo": motivo,
        }

    return _ok(response_data)


@app.post("/api/repartidor/asignaciones")
@login_required(roles=["admin"])
def api_asignar_pedido_repartidor():
    payload = request.get_json(silent=True) or {}
    pedido_id = payload.get("pedido_id")
    repartidor_usuario = payload.get("repartidor_usuario")

    if not pedido_id:
        return _error("pedido_id es obligatorio", 400)

    if not repartidor_usuario:
        return _error(
            "repartidor_usuario es obligatorio. Usa un usuario de rol repartidor con area_entrega configurada.",
            400,
        )

    created = db.asignar_pedido_repartidor(
        pedido_id=pedido_id,
        repartidor_usuario=repartidor_usuario,
        asignado_por=session.get("user", {}).get("username", "admin"),
    )
    if isinstance(created, dict) and created.get("error"):
        msg = created["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(created["error"], status)
    return _ok(created, 201)


@app.get("/api/pedidos/<int:pedido_id>/bitacora")
@login_required(roles=["admin", "cocina", "repartidor"])
def api_bitacora_pedido(pedido_id):
    limit = request.args.get("limit", "50")
    try:
        limit_int = max(1, min(200, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_bitacora_pedido(pedido_id=pedido_id, limit=limit_int)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.post("/api/pedidos/<int:pedido_id>/reenviar-codigo")
@login_required(roles=["admin", "repartidor"])
def api_reenviar_codigo_pedido(pedido_id):
    data = db.obtener_o_generar_codigo_entrega_pedido(pedido_id=pedido_id)
    if isinstance(data, dict) and data.get("error"):
        msg = data["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(data["error"], status)

    destino_data = db.obtener_destino_whatsapp_por_pedido(pedido_id=pedido_id)
    if isinstance(destino_data, dict) and destino_data.get("error"):
        msg = destino_data["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(destino_data["error"], status)

    codigo = data.get("codigo_entrega")
    destino = _normalizar_whatsapp_id(destino_data.get("whatsapp_id"))
    mensaje = (
        f"Buena nota, tu codigo de entrega para el pedido #{pedido_id} es: {codigo}. "
        "Compartelo al repartidor para liberar el pedido."
    )

    enviado = _enviar_texto_whatsapp(destino=destino, texto=mensaje)
    if isinstance(enviado, dict) and enviado.get("error"):
        return _error(f"No se pudo reenviar el codigo por WhatsApp: {enviado['error']}", 502)

    db.crear_log_notificacion(
        {
            "pedido_id": pedido_id,
            "canal": "whatsapp",
            "destino": destino,
            "tipo": "codigo_entrega",
            "mensaje": mensaje,
            "total": None,
            "direccion": None,
        }
    )

    return _ok(
        {
            "pedido_id": pedido_id,
            "codigo_entrega": codigo,
            "destino": destino,
            "message": "Codigo reenviado por WhatsApp al cliente.",
        }
    )


@app.post("/api/evaluaciones/programar")
@login_required(roles=["admin", "repartidor"])
def api_programar_evaluacion_entrega():
    payload = request.get_json(silent=True) or {}
    pedido_id = payload.get("pedido_id")
    retraso_minutos = payload.get("retraso_minutos", 15)
    if not pedido_id:
        return _error("pedido_id es obligatorio", 400)
    return _ok(
        {
            "pedido_id": pedido_id,
            "retraso_minutos": retraso_minutos,
            "message": "Evaluacion programada (demo).",
        }
    )


@app.get("/api/clientes/top20")
@login_required(roles=["admin"])
def api_top_clientes():
    top = db.obtener_top_clientes(limit=20)
    if isinstance(top, dict) and top.get("error"):
        return _error(top["error"], 500)
    return _ok(top)


@app.get("/api/ventas/diarias")
@login_required(roles=["admin"])
def api_ventas_diarias():
    data = db.obtener_ventas_diarias()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/ventas/mensuales")
@login_required(roles=["admin"])
def api_ventas_mensuales():
    data = db.obtener_ventas_mensuales()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/ventas/anuales")
@login_required(roles=["admin"])
def api_ventas_anuales():
    data = db.obtener_ventas_anuales()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/ventas/kpis-periodo")
@login_required(roles=["admin"])
def api_kpis_ventas_periodo():
    data = db.obtener_kpis_ventas_periodo()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/reporte-ventas-profesional")
@login_required(roles=["admin"])
def api_reporte_ventas_profesional():
    periodo = (request.args.get("periodo") or "dia").strip().lower()
    fecha_base = (request.args.get("fecha") or "").strip() or None
    busqueda = (request.args.get("q") or request.args.get("buscar") or "").strip() or None

    limit_raw = request.args.get("limit", "300")
    try:
        limit_int = max(1, min(1000, int(limit_raw)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_reporte_ventas_profesional(
        periodo=periodo,
        fecha_base=fecha_base,
        busqueda=busqueda,
        limit=limit_int,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/reporte-ventas-profesional.xlsx")
@login_required(roles=["admin"])
def api_reporte_ventas_profesional_xlsx():
    periodo = (request.args.get("periodo") or "dia").strip().lower()
    fecha_base = (request.args.get("fecha") or "").strip() or None
    busqueda = (request.args.get("q") or request.args.get("buscar") or "").strip() or None

    limit_raw = request.args.get("limit", "1000")
    try:
        limit_int = max(1, min(5000, int(limit_raw)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_reporte_ventas_profesional(
        periodo=periodo,
        fecha_base=fecha_base,
        busqueda=busqueda,
        limit=limit_int,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)

    try:
        from openpyxl import Workbook
    except Exception:
        return _error("Falta dependencia openpyxl. Ejecuta: pip install openpyxl", 500)

    payload = _serialize(data if isinstance(data, dict) else {})
    resumen = payload.get("resumen") or {}
    rows = payload.get("rows") or []

    wb = Workbook()
    ws_res = wb.active
    if ws_res is None:
        ws_res = wb.create_sheet(title="Resumen")
    else:
        ws_res.title = "Resumen"
    ws_res.append(["Metrica", "Valor"])
    periodo_label = {"dia": "Día", "semana": "Semana", "mes": "Mes", "ano": "Año"}.get(
        str(payload.get("periodo") or "dia").strip().lower(),
        str(payload.get("periodo") or "dia"),
    )
    ws_res.append(["Periodo", periodo_label])
    ws_res.append(["Fecha base", payload.get("fecha_base") or ""])
    ws_res.append(["Ventas", resumen.get("ventas") or 0])
    ws_res.append(["Pedidos", resumen.get("pedidos") or 0])
    ws_res.append(["Ticket promedio", resumen.get("ticket_promedio") or 0])
    ws_res.append(["Clientes unicos", resumen.get("clientes_unicos") or 0])
    ws_res.append(["Costo estimado total", resumen.get("costo_estimado_total") or 0])
    ws_res.append(["Utilidad estimada total", resumen.get("utilidad_estimada_total") or 0])
    ws_res.append(["Margen estimado pct", resumen.get("margen_estimado_pct") or 0])
    ws_res.append(["Rapidez preparacion promedio min", resumen.get("rapidez_preparacion_promedio_min") or 0])
    ws_res.append(["Rapidez entrega promedio min", resumen.get("rapidez_entrega_promedio_min") or 0])

    ws_det = wb.create_sheet(title="Detalle")
    ws_det.append([
        "pedido_id",
        "creado_en",
        "cliente",
        "whatsapp_id",
        "metodo_pago",
        "metodo_entrega",
        "estado",
        "productos",
        "piezas",
        "total",
        "costo_estimado",
        "utilidad_estimada",
        "margen_estimado_pct",
        "rapidez_preparacion_min",
        "rapidez_entrega_min",
    ])

    for row in rows:
        if not isinstance(row, dict):
            continue
        ws_det.append([
            row.get("pedido_id"),
            row.get("creado_en"),
            row.get("cliente"),
            row.get("whatsapp_id"),
            row.get("metodo_pago"),
            row.get("metodo_entrega"),
            row.get("estado"),
            row.get("productos"),
            row.get("piezas"),
            row.get("total"),
            row.get("costo_estimado"),
            row.get("utilidad_estimada"),
            row.get("margen_estimado_pct"),
            row.get("rapidez_preparacion_min"),
            row.get("rapidez_entrega_min"),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_ventas_profesional_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.get("/api/inventario/alertas")
@login_required(roles=["admin"])
def api_alertas_inventario():
    data = db.obtener_alertas_inventario()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/inventario")
@login_required(roles=["admin"])
def api_inventario():
    texto = (request.args.get("q") or request.args.get("texto") or "").strip() or None
    estado_stock = (request.args.get("estado_stock") or "").strip().lower() or None
    proveedor = (request.args.get("proveedor") or "").strip() or None

    limit_raw = request.args.get("limit")
    offset_raw = request.args.get("offset", "0")
    limit_int = None
    offset_int = 0
    if limit_raw not in (None, ""):
        try:
            limit_int = max(1, min(500, int(limit_raw)))
        except ValueError:
            return _error("Parametro limit invalido", 400)
    try:
        offset_int = max(0, int(offset_raw))
    except ValueError:
        return _error("Parametro offset invalido", 400)

    data = db.obtener_inventario(
        texto=texto,
        estado_stock=estado_stock,
        proveedor=proveedor,
        limit=limit_int,
        offset=offset_int,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.post("/api/inventario/compras")
@login_required(roles=["admin"])
def api_inventario_compras():
    payload = request.get_json(silent=True) or {}

    insumo = payload.get("insumo")
    cantidad = payload.get("cantidad")
    proveedor = payload.get("proveedor")
    costo_total = payload.get("costo_total")

    created = db.registrar_compra_insumo(
        insumo=insumo,
        cantidad=cantidad,
        proveedor=proveedor,
        costo_total=costo_total,
        creado_por=session.get("user", {}).get("username"),
        actor_rol=session.get("user", {}).get("rol", "admin"),
    )
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 400)
    return _ok(created, 201)


@app.get("/api/inventario/compras")
@login_required(roles=["admin"])
def api_historial_compras_inventario():
    limit = request.args.get("limit", "30")
    try:
        limit_int = max(1, min(200, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_compras_insumos(limit=limit_int)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/resumen-db")
@login_required(roles=["admin"])
def api_admin_resumen_db():
    data = db.obtener_resumen_db()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/rentabilidad-productos")
@login_required(roles=["admin"])
def api_admin_rentabilidad_productos():
    limit = request.args.get("limit", "20")
    try:
        limit_int = max(1, min(200, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_rentabilidad_productos(limit=limit_int)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/productos")
@login_required(roles=["admin"])
def api_admin_productos_listar():
    limit = request.args.get("limit", "200")
    try:
        limit_int = max(1, min(500, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_productos_admin(limit=limit_int)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.post("/api/admin/productos")
@login_required(roles=["admin"])
def api_admin_productos_crear_actualizar():
    payload = request.get_json(silent=True) or {}
    created = db.crear_producto_manual(
        nombre=payload.get("nombre"),
        variante=payload.get("variante"),
        precio=payload.get("precio"),
        activo=payload.get("activo", True),
    )
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 400)
    return _ok(created, 201)


@app.patch("/api/admin/productos/<int:producto_id>")
@login_required(roles=["admin"])
def api_admin_productos_actualizar(producto_id):
    payload = request.get_json(silent=True) or {}
    updated = db.actualizar_producto_admin(
        producto_id=producto_id,
        nombre=payload.get("nombre") if "nombre" in payload else None,
        variante=payload.get("variante") if "variante" in payload else None,
        precio=payload.get("precio") if "precio" in payload else None,
        activo=payload.get("activo") if "activo" in payload else None,
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated)


@app.post("/api/admin/insumos")
@login_required(roles=["admin"])
def api_admin_insumos_crear_actualizar():
    payload = request.get_json(silent=True) or {}
    actor = session.get("user", {})
    created = db.crear_insumo_manual(
        nombre=payload.get("nombre"),
        unidad_medida=payload.get("unidad_medida"),
        stock_minimo=payload.get("stock_minimo", 0),
        stock_inicial=payload.get("stock_inicial", 0),
        proveedor=payload.get("proveedor"),
        actor_username=actor.get("username"),
        actor_rol=actor.get("rol", "admin"),
    )
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 400)
    return _ok(created, 201)


@app.patch("/api/admin/insumos/<int:insumo_id>")
@login_required(roles=["admin"])
def api_admin_insumos_actualizar(insumo_id):
    payload = request.get_json(silent=True) or {}
    updated = db.actualizar_insumo_admin(
        insumo_id=insumo_id,
        unidad_medida=payload.get("unidad_medida") if "unidad_medida" in payload else None,
        stock_minimo=payload.get("stock_minimo") if "stock_minimo" in payload else None,
        proveedor=payload.get("proveedor") if "proveedor" in payload else None,
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated)


@app.post("/api/admin/insumos/<int:insumo_id>/ajuste-stock")
@login_required(roles=["admin"])
def api_admin_insumos_ajustar_stock(insumo_id):
    payload = request.get_json(silent=True) or {}
    actor = session.get("user", {})
    updated = db.ajustar_stock_insumo(
        insumo_id=insumo_id,
        cantidad_ajuste=payload.get("cantidad_ajuste"),
        motivo=payload.get("motivo"),
        actor_username=actor.get("username"),
        actor_rol=actor.get("rol", "admin"),
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated, 201)


@app.post("/api/admin/recetas-producto")
@login_required(roles=["admin"])
def api_admin_recetas_producto_guardar_componente():
    payload = request.get_json(silent=True) or {}
    saved = db.guardar_componente_receta(
        producto_id=payload.get("producto_id"),
        insumo_id=payload.get("insumo_id"),
        cantidad_por_unidad=payload.get("cantidad_por_unidad"),
        activo=payload.get("activo", True),
    )
    if isinstance(saved, dict) and saved.get("error"):
        return _error(saved["error"], 400)
    return _ok(saved, 201)


@app.get("/api/admin/recetas-producto")
@login_required(roles=["admin"])
def api_admin_recetas_producto_listar():
    producto_id = request.args.get("producto_id")
    texto = (request.args.get("q") or request.args.get("texto") or "").strip() or None
    activa_raw = (request.args.get("activa") or "").strip().lower()
    activa = None
    if activa_raw in {"1", "true", "yes", "on"}:
        activa = True
    elif activa_raw in {"0", "false", "no", "off"}:
        activa = False

    limit_raw = request.args.get("limit")
    offset_raw = request.args.get("offset", "0")
    limit_int = None
    offset_int = 0
    if limit_raw not in (None, ""):
        try:
            limit_int = max(1, min(500, int(limit_raw)))
        except ValueError:
            return _error("Parametro limit invalido", 400)
    try:
        offset_int = max(0, int(offset_raw))
    except ValueError:
        return _error("Parametro offset invalido", 400)

    pid = None
    if producto_id not in (None, ""):
        try:
            pid = int(producto_id)
        except ValueError:
            return _error("Parametro producto_id invalido", 400)

    data = db.obtener_recetas_producto(
        producto_id=pid,
        texto=texto,
        activa=activa,
        limit=limit_int,
        offset=offset_int,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.patch("/api/admin/recetas-producto/<int:receta_id>")
@login_required(roles=["admin"])
def api_admin_recetas_producto_actualizar(receta_id):
    payload = request.get_json(silent=True) or {}
    updated = db.actualizar_componente_receta(
        receta_id=receta_id,
        activo=payload.get("activo") if "activo" in payload else None,
        cantidad_por_unidad=payload.get("cantidad_por_unidad") if "cantidad_por_unidad" in payload else None,
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated)


@app.get("/api/admin/inventario/movimientos")
@login_required(roles=["admin"])
def api_admin_inventario_movimientos():
    limit = request.args.get("limit", "50")
    insumo_id = request.args.get("insumo_id")
    tipo = (request.args.get("tipo") or "").strip() or None

    try:
        limit_int = max(1, min(500, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    iid = None
    if insumo_id not in (None, ""):
        try:
            iid = int(insumo_id)
        except ValueError:
            return _error("Parametro insumo_id invalido", 400)

    data = db.obtener_movimientos_inventario(limit=limit_int, insumo_id=iid, tipo=tipo)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/productos-sin-receta")
@login_required(roles=["admin"])
def api_admin_productos_sin_receta():
    data = db.obtener_productos_sin_receta()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/usuarios")
@login_required(roles=["admin"])
def api_admin_usuarios_listar():
    rol = (request.args.get("rol") or "").strip() or None
    area_entrega = (request.args.get("area_entrega") or "").strip() or None

    data = db.obtener_usuarios_sistema(rol=rol, area_entrega=area_entrega)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.post("/api/admin/usuarios")
@login_required(roles=["admin"])
def api_admin_usuarios_crear():
    payload = request.get_json(silent=True) or {}
    actor = session.get("user", {})
    username = payload.get("username")
    password = payload.get("password")
    rol = payload.get("rol")
    nombre_mostrar = payload.get("nombre_mostrar") or payload.get("nombre")
    telefono = payload.get("telefono")
    area_entrega = payload.get("area_entrega")

    created = db.crear_usuario_sistema(
        username=username,
        password=password,
        rol=rol,
        nombre_mostrar=nombre_mostrar,
        telefono=telefono,
        area_entrega=area_entrega,
        actor_usuario_id=actor.get("usuario_id"),
        actor_username=actor.get("username"),
        actor_rol=actor.get("rol"),
        direccion_ip=_client_ip(),
    )
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 400)
    return _ok(created, 201)


@app.patch("/api/admin/usuarios/<int:usuario_id>")
@login_required(roles=["admin"])
def api_admin_usuarios_actualizar(usuario_id):
    payload = request.get_json(silent=True) or {}
    actor = session.get("user", {})
    updated = db.actualizar_usuario_sistema(
        usuario_id=usuario_id,
        rol=payload.get("rol"),
        nombre_mostrar=payload.get("nombre_mostrar"),
        telefono=payload.get("telefono"),
        area_entrega=payload.get("area_entrega") if "area_entrega" in payload else None,
        activo=payload.get("activo") if "activo" in payload else None,
        nueva_password=payload.get("nueva_password"),
        actor_usuario_id=actor.get("usuario_id"),
        actor_username=actor.get("username"),
        actor_rol=actor.get("rol"),
        direccion_ip=_client_ip(),
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated)


@app.get("/api/admin/auditoria-seguridad")
@login_required(roles=["admin"])
def api_admin_auditoria_seguridad():
    limit = request.args.get("limit", "40")
    offset = request.args.get("offset", "0")
    tipo_evento = (request.args.get("tipo_evento") or "").strip() or None
    severidad = (request.args.get("severidad") or "").strip().lower() or None
    actor = (request.args.get("actor") or "").strip() or None
    fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
    rango_rapido = (request.args.get("rango") or request.args.get("rango_rapido") or "").strip().lower() or None

    severidades_validas = {"info", "warning", "critical"}
    if severidad and severidad not in severidades_validas:
        return _error("Parametro severidad invalido", 400)

    try:
        limit_int = max(1, min(200, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)
    try:
        offset_int = max(0, int(offset))
    except ValueError:
        return _error("Parametro offset invalido", 400)

    data = db.obtener_auditoria_seguridad(
        limit=limit_int,
        offset=offset_int,
        tipo_evento=tipo_evento,
        severidad=severidad,
        actor_username=actor,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        rango_rapido=rango_rapido,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/auditoria-seguridad.csv")
@login_required(roles=["admin"])
def api_admin_auditoria_seguridad_csv():
    tipo_evento = (request.args.get("tipo_evento") or "").strip() or None
    severidad = (request.args.get("severidad") or "").strip().lower() or None
    actor = (request.args.get("actor") or "").strip() or None
    fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
    limit = request.args.get("limit", "1000")

    severidades_validas = {"info", "warning", "critical"}
    if severidad and severidad not in severidades_validas:
        return _error("Parametro severidad invalido", 400)

    try:
        limit_int = max(1, min(5000, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_auditoria_seguridad(
        limit=limit_int,
        tipo_evento=tipo_evento,
        severidad=severidad,
        actor_username=actor,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "auditoria_id",
        "fecha",
        "tipo_evento",
        "severidad",
        "actor_usuario_id",
        "actor_username",
        "actor_rol",
        "objetivo_usuario_id",
        "objetivo_username",
        "direccion_ip",
        "detalle",
    ])

    for row in data:
        if not isinstance(row, dict):
            continue
        writer.writerow([
            row.get("auditoria_id"),
            row.get("creado_en"),
            row.get("tipo_evento"),
            row.get("severidad"),
            row.get("actor_usuario_id"),
            row.get("actor_username"),
            row.get("actor_rol"),
            row.get("objetivo_usuario_id"),
            row.get("objetivo_username"),
            row.get("direccion_ip"),
            json.dumps(row.get("detalle") or {}, ensure_ascii=False),
        ])

    filename = f"auditoria_seguridad_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.get("/api/admin/auditoria-negocio")
@login_required(roles=["admin"])
def api_admin_auditoria_negocio():
    limit = request.args.get("limit", "40")
    offset = request.args.get("offset", "0")
    tabla = (request.args.get("tabla") or "").strip().lower() or None
    actor = (request.args.get("actor") or "").strip() or None
    fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
    rango_rapido = (request.args.get("rango") or request.args.get("rango_rapido") or "").strip().lower() or None

    tablas_validas = {"pedidos", "pagos", "insumos", "compras_insumos"}
    if tabla and tabla not in tablas_validas:
        return _error("Parametro tabla invalido", 400)

    try:
        limit_int = max(1, min(200, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)
    try:
        offset_int = max(0, int(offset))
    except ValueError:
        return _error("Parametro offset invalido", 400)

    data = db.obtener_auditoria_negocio(
        limit=limit_int,
        offset=offset_int,
        tabla_objetivo=tabla,
        actor_username=actor,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        rango_rapido=rango_rapido,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/admin/auditoria-negocio.csv")
@login_required(roles=["admin"])
def api_admin_auditoria_negocio_csv():
    tabla = (request.args.get("tabla") or "").strip().lower() or None
    actor = (request.args.get("actor") or "").strip() or None
    fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
    limit = request.args.get("limit", "1000")

    tablas_validas = {"pedidos", "pagos", "insumos", "compras_insumos"}
    if tabla and tabla not in tablas_validas:
        return _error("Parametro tabla invalido", 400)

    try:
        limit_int = max(1, min(5000, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_auditoria_negocio(
        limit=limit_int,
        tabla_objetivo=tabla,
        actor_username=actor,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "auditoria_negocio_id",
        "fecha",
        "tabla_objetivo",
        "operacion",
        "registro_id",
        "actor_username",
        "actor_rol",
        "detalle",
    ])

    for row in data:
        if not isinstance(row, dict):
            continue
        writer.writerow([
            row.get("auditoria_negocio_id"),
            row.get("creado_en"),
            row.get("tabla_objetivo"),
            row.get("operacion"),
            row.get("registro_id"),
            row.get("actor_username"),
            row.get("actor_rol"),
            json.dumps(row.get("detalle") or {}, ensure_ascii=False),
        ])

    filename = f"auditoria_negocio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.get("/api/admin/parser/observaciones")
@login_required(roles=["admin"])
def api_admin_parser_observaciones():
    limit = request.args.get("limit", "80")
    tipo_evento = (request.args.get("tipo_evento") or "").strip() or None
    estado_revision = (request.args.get("estado_revision") or "").strip().lower() or None
    try:
        limit_int = max(1, min(300, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)
    data = db.obtener_observaciones_parser(limit=limit_int, tipo_evento=tipo_evento, estado_revision=estado_revision)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.patch("/api/admin/parser/observaciones/<int:observacion_id>")
@login_required(roles=["admin"])
def api_admin_parser_observacion_actualizar(observacion_id):
    payload = request.get_json(silent=True) or {}
    try:
        expected_items = _parse_json_field(payload.get("expected_items_json"), None)
    except Exception:
        return _error("expected_items_json debe ser JSON valido", 400)
    updated = db.actualizar_observacion_parser(
        observacion_id=observacion_id,
        estado_revision=payload.get("estado_revision") if "estado_revision" in payload else None,
        admin_notes=payload.get("admin_notes") if "admin_notes" in payload else None,
        expected_items_json=expected_items,
        regla_id=payload.get("regla_id") if "regla_id" in payload else None,
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrada" in msg or "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated)


@app.get("/api/admin/parser/frases")
@login_required(roles=["admin"])
def api_admin_parser_frases():
    limit = request.args.get("limit", "200")
    activa_raw = (request.args.get("activa") or "").strip().lower()
    activa = None
    if activa_raw in {"1", "true", "yes", "on"}:
        activa = True
    elif activa_raw in {"0", "false", "no", "off"}:
        activa = False
    try:
        limit_int = max(1, min(500, int(limit)))
    except ValueError:
        return _error("Parametro limit invalido", 400)
    data = db.obtener_frases_parser_curadas(limit=limit_int, activa=activa)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.post("/api/admin/parser/frases")
@login_required(roles=["admin"])
def api_admin_parser_frases_crear():
    payload = request.get_json(silent=True) or {}
    try:
        items_json = _parse_json_field(payload.get("items_json"), [])
    except Exception:
        return _error("items_json debe ser JSON valido", 400)
    created = db.crear_frase_parser_curada(
        frase_original=payload.get("frase_original"),
        tipo_match=payload.get("tipo_match", "exact"),
        items_json=items_json,
        needs_confirmation=payload.get("needs_confirmation", False),
        needs_clarification=payload.get("needs_clarification", False),
        clarification_message=payload.get("clarification_message"),
        notas=payload.get("notas"),
        prioridad=payload.get("prioridad", 100),
        activa=payload.get("activa", True),
    )
    if isinstance(created, dict) and created.get("error"):
        return _error(created["error"], 400)
    return _ok(created, 201)


@app.patch("/api/admin/parser/frases/<int:regla_id>")
@login_required(roles=["admin"])
def api_admin_parser_frases_actualizar(regla_id):
    payload = request.get_json(silent=True) or {}
    try:
        items_json = _parse_json_field(payload.get("items_json"), None) if "items_json" in payload else None
    except Exception:
        return _error("items_json debe ser JSON valido", 400)
    updated = db.actualizar_frase_parser_curada(
        regla_id=regla_id,
        frase_original=payload.get("frase_original") if "frase_original" in payload else None,
        tipo_match=payload.get("tipo_match") if "tipo_match" in payload else None,
        items_json=items_json,
        needs_confirmation=payload.get("needs_confirmation") if "needs_confirmation" in payload else None,
        needs_clarification=payload.get("needs_clarification") if "needs_clarification" in payload else None,
        clarification_message=payload.get("clarification_message") if "clarification_message" in payload else None,
        notas=payload.get("notas") if "notas" in payload else None,
        prioridad=payload.get("prioridad") if "prioridad" in payload else None,
        activa=payload.get("activa") if "activa" in payload else None,
    )
    if isinstance(updated, dict) and updated.get("error"):
        msg = updated["error"].lower()
        status = 404 if "no encontrada" in msg or "no encontrado" in msg else 400
        return _error(updated["error"], status)
    return _ok(updated)


@app.post("/api/admin/parser/simular")
@login_required(roles=["admin"])
def api_admin_parser_simular():
    payload = request.get_json(silent=True) or {}
    texto = (payload.get("texto") or "").strip()
    if not texto:
        return _error("texto es obligatorio", 400)
    extractor = getattr(_bot_module, "_extraer_items_menu_oficial", None) if _bot_module is not None else None
    formatter = getattr(_bot_module, "_formatear_carrito", None) if _bot_module is not None else None
    if not callable(extractor):
        return _error("Extractor del bot no disponible", 500)
    try:
        extraccion = extractor(texto)
        resumen = ""
        if callable(formatter):
            items = extraccion.get("items") or []
            total = int(extraccion.get("total") or 0)
            resumen = formatter(items, total) if items else ""
        return _ok({"extraccion": extraccion, "resumen": resumen})
    except Exception as exc:
        return _error(f"No se pudo simular el parser: {exc}", 500)


@app.post("/api/campanias")
@login_required(roles=["admin"])
def api_campanias():
    payload = request.get_json(silent=True) or {}
    nombre = payload.get("nombre")
    mensaje = payload.get("mensaje")
    segmento = payload.get("segmento", "general")

    if not nombre or not mensaje:
        return _error("Los campos nombre y mensaje son obligatorios")

    creada = db.crear_campania(
        nombre=nombre,
        mensaje=mensaje,
        segmento=segmento,
        creada_por=session.get("user", {}).get("username"),
    )
    if isinstance(creada, dict) and creada.get("error"):
        return _error(creada["error"], 500)

    clientes = db.obtener_clientes_para_campania(filtro=segmento)
    if isinstance(clientes, dict) and clientes.get("error"):
        return _error(clientes["error"], 500)

    lista_clientes = clientes if isinstance(clientes, list) else []
    enviados = 0
    fallidos = 0

    for cliente in lista_clientes:
        if not isinstance(cliente, dict):
            continue

        destino = _normalizar_whatsapp_id(cliente.get("whatsapp_id"))
        if not destino:
            fallidos += 1
            db.registrar_envio_campana(
                campana_id=creada.get("campana_id"),
                cliente_id=cliente.get("cliente_id"),
                whatsapp_id="",
                enviado=False,
                error="Cliente sin whatsapp_id valido",
            )
            continue

        envio = _enviar_texto_whatsapp(destino=destino, texto=mensaje)
        if isinstance(envio, dict) and envio.get("error"):
            fallidos += 1
            db.registrar_envio_campana(
                campana_id=creada.get("campana_id"),
                cliente_id=cliente.get("cliente_id"),
                whatsapp_id=destino,
                enviado=False,
                error=envio.get("error"),
            )
            continue

        enviados += 1
        db.registrar_envio_campana(
            campana_id=creada.get("campana_id"),
            cliente_id=cliente.get("cliente_id"),
            whatsapp_id=destino,
            enviado=True,
            error=None,
        )

    respuesta = dict(creada)
    respuesta["total_destinatarios"] = len(lista_clientes)
    respuesta["mensajes_enviados"] = enviados
    respuesta["mensajes_fallidos"] = fallidos
    return _ok(respuesta, 201)


@app.get("/api/campanias")
@login_required(roles=["admin"])
def api_campanias_historial():
    limit_raw = request.args.get("limit", "80")
    try:
        limit = max(1, min(500, int(limit_raw)))
    except ValueError:
        return _error("Parametro limit invalido", 400)

    data = db.obtener_campanias(limit=limit)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/campanias/historial")
@login_required(roles=["admin"])
def api_campanias_historial_alias():
    return api_campanias_historial()


@app.get("/api/clientes/count")
@login_required(roles=["admin"])
def api_clientes_count():
    filtro = (request.args.get("filtro") or "todos").strip().lower()
    data = db.contar_clientes_para_campania(filtro=filtro)
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


@app.get("/api/empleados")
@login_required(roles=["admin"])
def api_empleados():
    data = db.obtener_empleados()
    if isinstance(data, dict) and data.get("error"):
        return _error(data["error"], 500)
    return _ok(data)


# ===================== SOPORTE / TICKETS =====================

@app.post("/api/soporte/tickets")
def api_crear_ticket():
    """Crea un ticket de soporte. Ruta publica, no requiere sesion."""
    payload = request.get_json(silent=True) or {}
    nombre = (payload.get("nombre_contacto") or "").strip()
    descripcion = (payload.get("descripcion") or "").strip()
    if not nombre:
        return _error("nombre_contacto es obligatorio", 400, code="validation_error")
    if not descripcion:
        return _error("descripcion es obligatoria", 400, code="validation_error")
    result = db.crear_ticket_soporte(
        categoria=(payload.get("categoria") or "otro").strip().lower(),
        prioridad=(payload.get("prioridad") or "media").strip().lower(),
        nombre_contacto=nombre,
        whatsapp_contacto=(payload.get("whatsapp_contacto") or "").strip() or None,
        descripcion=descripcion,
    )
    if isinstance(result, dict) and result.get("error"):
        return _error(result["error"], 500)
    return _ok(result, status=201)


@app.get("/api/soporte/tickets")
@login_required(roles=["admin"])
def api_listar_tickets():
    estado = request.args.get("estado") or None
    numero = (request.args.get("numero") or "").strip().upper()
    result = db.obtener_tickets_soporte(estado=estado)
    if isinstance(result, dict) and result.get("error"):
        return _error(result["error"], 500)
    if numero:
        data = [t for t in result if isinstance(t, dict) and (t.get("numero_ticket") or "").strip().upper() == numero]
        return _ok(data)
    return _ok(result)


@app.get("/api/soporte/tickets/public/<numero>")
def api_consultar_ticket_publico(numero):
    numero_norm = (numero or "").strip().upper()
    if not numero_norm:
        return _error("numero de ticket invalido", 400, code="validation_error")

    result = db.obtener_tickets_soporte(estado=None)
    if isinstance(result, dict) and result.get("error"):
        return _error(result["error"], 500)

    ticket = None
    for item in result:
        if isinstance(item, dict) and (item.get("numero_ticket") or "").strip().upper() == numero_norm:
            ticket = item
            break

    if not ticket:
        return _error(f"Ticket {numero_norm} no encontrado", 404, code="ticket_not_found")

    publico = {
        "numero_ticket": ticket.get("numero_ticket"),
        "categoria": ticket.get("categoria"),
        "prioridad": ticket.get("prioridad"),
        "estado": ticket.get("estado"),
        "descripcion": ticket.get("descripcion"),
        "creado_en": ticket.get("creado_en"),
        "actualizado_en": ticket.get("actualizado_en"),
        "notas_resolucion": ticket.get("notas_resolucion"),
    }
    return _ok(publico)


@app.patch("/api/soporte/tickets/<numero>")
@login_required(roles=["admin"])
def api_actualizar_ticket(numero):
    payload = request.get_json(silent=True) or {}
    nuevo_estado = (payload.get("estado") or "").strip()
    if not nuevo_estado:
        return _error("estado es obligatorio", 400, code="validation_error")
    actor = cast(Any, session).get("user", {}).get("username")
    result = db.actualizar_estado_ticket(
        numero_ticket=numero,
        nuevo_estado=nuevo_estado,
        notas_resolucion=(payload.get("notas_resolucion") or "").strip() or None,
        resuelto_por=actor,
    )
    if isinstance(result, dict) and result.get("error"):
        return _error(result["error"], 400)

    notificacion = {
        "intentada": False,
        "enviada": False,
        "destino": None,
        "motivo": "No aplica para este estado.",
    }
    estado_final = (result.get("estado") or "").strip().lower()
    if estado_final in {"resuelto", "cerrado"}:
        notificacion["intentada"] = True
        destino = _normalizar_destino_ticket_whatsapp(result.get("whatsapp_contacto"))
        notificacion["destino"] = destino or None
        if destino:
            notas = (result.get("notas_resolucion") or "").strip()
            texto = (
                f"Listo parce, tu ticket {result.get('numero_ticket')} ya quedo {estado_final}. "
                "Gracias por reportarlo."
            )
            if notas:
                texto = f"{texto} Nota de soporte: {notas}"
            enviado = _enviar_texto_whatsapp(destino=destino, texto=texto)
            if isinstance(enviado, dict) and enviado.get("error"):
                notificacion["motivo"] = enviado["error"]
            else:
                notificacion["enviada"] = True
                notificacion["motivo"] = "ok"
        else:
            notificacion["motivo"] = "Ticket sin whatsapp_contacto valido."

    response_data = dict(result)
    response_data["notificacion_whatsapp"] = notificacion
    return _ok(response_data)


@app.post("/webhook")
def webhook_whatsapp():
    whatsapp_id = _normalizar_whatsapp_id(request.values.get("From", ""))
    result = procesar_mensaje_whatsapp(
        whatsapp_id=whatsapp_id,
        mensaje=request.values.get("Body", ""),
        media_url=request.values.get("MediaUrl0"),
        media_type=request.values.get("MediaContentType0"),
        latitude=request.values.get("Latitude"),
        longitude=request.values.get("Longitude"),
    )
    if not isinstance(result, dict):
        result = {"tipo": "texto", "contenido": str(result)}

    twiml = MessagingResponse()
    msg = twiml.message()

    tipo = result.get("tipo", "texto")
    if tipo == "audio" and result.get("audio_filename"):
        base_url = app.config["PUBLIC_BASE_URL"] or request.url_root.rstrip("/")
        audio_url = f"{base_url}/audio/{result['audio_filename']}"
        msg.media(audio_url)
        if result.get("contenido"):
            msg.body(result["contenido"])
    else:
        msg.body(result.get("contenido", "Listo parce, mensaje recibido."))

    return str(twiml), 200, {"Content-Type": "text/xml; charset=utf-8"}


@app.post("/webhook/baileys")
def webhook_baileys():
    expected_token = app.config.get("BAILEYS_WEBHOOK_TOKEN", "")
    if expected_token:
        incoming_token = (request.headers.get("x-bridge-token") or "").strip()
        if incoming_token != expected_token:
            return _error("Token de bridge invalido", 401)

    payload = request.get_json(silent=True) or {}

    whatsapp_id = _normalizar_whatsapp_id(payload.get("whatsapp_id") or payload.get("from") or payload.get("jid"))
    whatsapp_jid = (payload.get("whatsapp_jid") or payload.get("jid") or "").strip()
    mensaje = payload.get("mensaje") or payload.get("text") or ""
    media_url = payload.get("media_url") or payload.get("mediaUrl")
    media_type = payload.get("media_type") or payload.get("mediaType")
    media_kind = payload.get("media_kind") or payload.get("mediaKind")
    latitude = payload.get("latitude")
    longitude = payload.get("longitude")

    app.logger.info(
        "Webhook Baileys recibido: whatsapp_id=%s has_text=%s media_type=%s has_media_url=%s",
        whatsapp_id or "",
        bool(str(mensaje or "").strip()),
        media_type or "",
        bool(media_url),
    )

    if not whatsapp_id:
        return _error("whatsapp_id es obligatorio", 400)

    output = procesar_mensaje_whatsapp(
        whatsapp_id=whatsapp_id,
        mensaje=mensaje,
        media_url=media_url,
        media_type=media_type,
        media_kind=media_kind,
        latitude=latitude,
        longitude=longitude,
    )
    if not isinstance(output, dict):
        output = {"tipo": "texto", "contenido": str(output)}

    if output.get("tipo") == "audio" and output.get("audio_filename") and not output.get("audio_url"):
        base_url = app.config["PUBLIC_BASE_URL"] or request.url_root.rstrip("/")
        output["audio_url"] = f"{base_url}/audio/{output['audio_filename']}"

    # Si el dispatcher genero audio colombiano de transicion de estado, enviarlo
    # como nota de voz adicional en segundo plano (fire-and-forget).
    audio_colombiano = output.get("audio_colombiano_path")
    should_send_transition_audio = bool(audio_colombiano) and bool(whatsapp_id) and output.get("tipo") != "audio"
    if should_send_transition_audio:
        import threading
        destino_audio = whatsapp_jid or whatsapp_id
        def _enviar_colombiano_bg():
            try:
                _enviar_audio_whatsapp(destino_audio, audio_colombiano)
            except Exception as _exc:
                app.logger.debug("audio_colombiano bg error: %s", _exc)
        threading.Thread(target=_enviar_colombiano_bg, daemon=True).start()

    app.logger.info(
        "Webhook Baileys respuesta: whatsapp_id=%s tipo=%s has_audio_url=%s has_audio_colombiano=%s",
        whatsapp_id,
        output.get("tipo", "texto"),
        bool(output.get("audio_url")),
        bool(audio_colombiano),
    )

    return _ok(output)


def _render_error_page(status_code, title, message, error_id):
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    response = make_response(
        render_template(
            "error.html",
            status_code=status_code,
            title=title,
            message=message,
            error_id=error_id,
            timestamp=now_text,
        ),
        status_code,
    )
    response.headers["X-Error-ID"] = error_id
    return response


@app.errorhandler(400)
def handle_400(err):
    if _espera_json():
        return _error("Solicitud invalida", 400, code="bad_request")
    error_id = _generar_error_id()
    return _render_error_page(400, "Solicitud invalida", "La solicitud no pudo ser procesada. Verifica los datos enviados.", error_id)


@app.errorhandler(401)
def handle_401(err):
    if _espera_json():
        return _error("Sesion requerida", 401, code="auth_required")
    return redirect(url_for("login_page", next=request.path))


@app.errorhandler(403)
def handle_403(err):
    if _espera_json():
        return _error("No autorizado para esta accion", 403, code="forbidden")
    error_id = _generar_error_id()
    return _render_error_page(403, "Acceso denegado", "No tienes permisos para acceder a esta seccion.", error_id)


@app.errorhandler(404)
def handle_404(err):
    if _espera_json():
        return _error("Recurso no encontrado", 404, code="not_found")
    error_id = _generar_error_id()
    return _render_error_page(404, "Pagina no encontrada", "La pagina solicitada no existe o fue movida.", error_id)


@app.errorhandler(500)
def handle_500(err):
    error_id = _generar_error_id()
    app.logger.exception("Error interno no controlado [%s]: %s", error_id, err)
    if _espera_json():
        return _error("Error interno del servidor", 500, code="internal_error", error_id=error_id)
    return _render_error_page(500, "Error interno", "Ocurrio un problema inesperado. Intenta nuevamente en unos minutos.", error_id)


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug_enabled = (os.getenv("FLASK_DEBUG", "0") or "0").strip().lower() in {"1", "true", "yes", "on"}
    app.run(host="0.0.0.0", port=port, debug=debug_enabled)
